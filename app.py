import json
import os
import re
import uuid
import secrets
import sqlite3
from datetime import datetime, timezone, timedelta
from flask import Flask, render_template, request, redirect, url_for, flash, send_file, make_response, session, g, send_from_directory
from functools import wraps
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
import subprocess
import sys

app = Flask(__name__)
# Use environment variable for the secret key in production


def ensure_package(pkg_name, import_name=None):
    """Try to import a package; if missing, attempt to install it via pip then import.
    Returns the imported module or raises ImportError.
    """
    import importlib
    name = import_name or pkg_name
    try:
        return importlib.import_module(name)
    except ImportError:
        try:
            subprocess.check_call([sys.executable, '-m', 'pip', 'install', pkg_name])
        except Exception:
            raise ImportError(f"Não foi possível instalar {pkg_name}. Instale manualmente.")
        return importlib.import_module(name)
app.secret_key = os.environ.get("APP_SECRET", "troque-esta-chave-em-producao")

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
# No Vercel (serverless) o disco do projeto é só-leitura; gravamos em /tmp lá.
DATA_DIR = "/tmp" if os.environ.get("VERCEL") else BASE_DIR + "/data"
DB_PATH = os.path.join(DATA_DIR, "data.db")
USE_SQLITE = os.environ.get("USE_SQLITE", "1") in ("1", "true", "yes")
DATA_FILE = os.path.join(DATA_DIR, "materiais.json")
SEED_FILE = os.path.join(BASE_DIR, "data", "materiais.json")

CATEGORIAS_EMOJI = {
    "Courino": "🟫",
    "Metal": "⚙️",
    "Aviamento": "🧵",
    "Tecido": "🧶",
    "Embalagem": "📦",
    "Outros": "🔹",
}
CATEGORIAS = list(CATEGORIAS_EMOJI.keys())
UNIDADES = ["unidades", "metros", "rolos", "kg", "gramas", "pares", "pacotes"]
MOTIVOS_BAIXA = ["Produção de bolsa", "Produção de nécessaire", "Amostra / Teste", "Desperdício"]

EMOJIS_PRODUTO = ["👜", "🎒", "👝", "💼", "🧳", "👛"]
STATUS_PEDIDO = ["Pendente", "Em produção", "Concluído", "Entregue", "Cancelado"]
STATUS_PEDIDO_BADGE = {
    "Pendente": "badge-warn",
    "Em produção": "badge-warn",
    "Concluído": "badge-ok",
    "Entregue": "badge-ok",
    "Cancelado": "badge-low",
}
STATUS_SOBRA_BADGE = {
    "Disponível": "badge-warn",
    "Reaproveitado": "badge-ok",
    "Descartado": "badge-low",
}
CATEGORIAS_DESPESA = ["Matéria-prima", "Aluguel", "Transporte", "Ferramentas", "Marketing", "Outros"]


# ── Fuso horário (datas corretas no horário de Brasília) ─────────────────────

def _fuso_brasil():
    try:
        from zoneinfo import ZoneInfo
        return ZoneInfo("America/Sao_Paulo")
    except Exception:
        return timezone(timedelta(hours=-3))


FUSO_BR = _fuso_brasil()


def agora():
    """Datetime atual no fuso horário do Brasil (UTC-3)."""
    return datetime.now(FUSO_BR)


def formatar_reais(valor):
    """Formata um valor monetário no padrão brasileiro (ex.: 1.234,56)."""
    try:
        v = float(valor or 0)
    except (TypeError, ValueError):
        v = 0.0
    inteiro, decimal = f"{v:,.2f}".split(".")
    return inteiro.replace(",", ".") + "," + decimal


def parse_float_ptbr(valor_str, default=0.0):
    """Converte valores monetários ou numéricos com vírgula/ponto para float."""
    if valor_str is None:
        return default
    if isinstance(valor_str, (int, float)):
        return float(valor_str)
    s = str(valor_str).replace("R$", "").replace(" ", "").strip()
    if not s:
        return default
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except (ValueError, TypeError):
        return default


@app.template_filter("moeda")
def filtro_moeda(valor):
    return "R$ " + formatar_reais(valor)


# ── Categorias consistentes e sincronizadas ───────────────────────────────────

CATEGORIA_REGRA = [
    ("Courino", ["courino", "couro", "suede", "camurca", "camurça", "napa", "sintetico", "sintético"]),
    ("Metal", ["metal", "mosquetao", "mosquetão", "fivela", "argola", "corrente", "ilhos", "ilhós", "gancho", "rebite", "tachinha", "alca de metal", "alça de metal", "fecho de metal", "fecho"]),
    ("Aviamento", ["ziper", "zíper", "linha", "fio", "botao", "botão", "colchete", "elastico", "elástico", "velcro", "aviamento", "travado", "etiqueta", "laco", "laço", "renda", "passante", "gancheira", "fita", "viés", "vies"]),
    ("Tecido", ["tecido", "algodao", "algodão", "linho", "tricoline", "voal", "crepe", "sarja", "seda", "forro", "pano", "gabardine", "oxford", "jeans", "malha", "feltro", "lona"]),
    ("Embalagem", ["embalagem", "saco", "caixa", "papel", "plastico", "plástico", "polimero", "polímero", "adesivo", "fita adesiva", "tag", "sacola"]),
    ("Outros", ["tesoura", "alicate", "regua", "régua", "ferramenta", "cola", "giz", "agulha"]),
]


def sugerir_categoria(nome):
    """Sugere uma categoria canônica a partir do nome do material."""
    nome_l = (nome or "").lower()
    for cat, palavras in CATEGORIA_REGRA:
        for p in palavras:
            if p in nome_l:
                return cat
    return "Outros"


def normalizar_categoria(categoria, nome=None):
    """Mapeia categorias livres/despadronizadas para as categorias canônicas."""
    if categoria in CATEGORIAS:
        return categoria
    if nome and sugerir_categoria(nome) != "Outros":
        return sugerir_categoria(nome)
    if categoria and sugerir_categoria(categoria) != "Outros":
        return sugerir_categoria(categoria)
    return "Outros"


# ── Múltiplos papéis por usuário ──────────────────────────────────────────────

def serializar_roles(roles):
    """Converte a lista de papéis em texto JSON para persistir na coluna 'roles'."""
    if not roles:
        return ""
    if isinstance(roles, str):
        return roles
    return json.dumps([r for r in roles if r], ensure_ascii=False)


def usuario_roles_lista(user):
    """Retorna a lista de papéis do usuário (coluna 'roles' JSON; fallback para 'role')."""
    if not user:
        return []
    roles = user.get("roles")
    if isinstance(roles, list):
        return [r for r in roles if r]
    if isinstance(roles, str) and roles.strip():
        try:
            parsed = json.loads(roles)
            if isinstance(parsed, list):
                return [r for r in parsed if r]
        except Exception:
            pass
        return [r.strip() for r in re.sub(r'[\["\]\s]', "", roles).split(",") if r.strip()]
    single = user.get("role") or ""
    return [single] if single else []


@app.context_processor
def injetar_helpers_templates():
    """Disponibiliza funções utilitárias para uso direto nos templates."""
    return dict(
        usuario_roles_lista=usuario_roles_lista,
        formatar_reais=formatar_reais,
    )


# ── Persistência genérica (usada pelos módulos novos) ────────────────────────

# ── Definição Canônica de Abas e Módulos do Sistema ──────────────────────────

SYSTEM_TABS = [
    {
        "id": "estoque",
        "name": "Estoque",
        "emoji": "📦",
        "desc": "Consulta de materiais em estoque, custos e quantidades",
        "actions": ["read", "update", "delete"],
        "route": "estoque",
    },
    {
        "id": "adicionar",
        "name": "Adicionar Material",
        "emoji": "➕",
        "desc": "Cadastro de novos materiais e insumos",
        "actions": ["create"],
        "route": "adicionar",
    },
    {
        "id": "baixa",
        "name": "Dar Baixa",
        "emoji": "✂️",
        "desc": "Registro de saídas manuais e consumo de insumos",
        "actions": ["create"],
        "route": "baixa",
    },
    {
        "id": "produtos",
        "name": "Produtos & Receitas",
        "emoji": "👜",
        "desc": "Catálogo de produtos artesanais e receitas técnicas",
        "actions": ["read", "create", "delete"],
        "route": "produtos",
    },
    {
        "id": "pedidos",
        "name": "Pedidos dos Clientes",
        "emoji": "🧾",
        "desc": "Gestão de pedidos, cálculo de insumos e status",
        "actions": ["read", "create", "update", "delete"],
        "route": "pedidos",
    },
    {
        "id": "sobras",
        "name": "Sobras e Reaproveitamento",
        "emoji": "♻️",
        "desc": "Controle de retalhos, reaproveitamento e descarte",
        "actions": ["read", "create", "update", "delete"],
        "route": "sobras",
    },
    {
        "id": "financeiro",
        "name": "Financeiro",
        "emoji": "💰",
        "desc": "Fluxo de caixa, receitas e despesas",
        "actions": ["read", "create", "delete"],
        "route": "financeiro",
    },
    {
        "id": "relatorios",
        "name": "Alertas e Relatórios",
        "emoji": "📊",
        "desc": "Alertas de estoque mínimo, movimentações e exportações",
        "actions": ["read"],
        "route": "alertas",
    },
    {
        "id": "usuarios",
        "name": "Usuários",
        "emoji": "👥",
        "desc": "Gestão de contas e senhas de usuários",
        "actions": ["read", "create", "update", "delete"],
        "route": "usuarios",
    },
    {
        "id": "roles",
        "name": "Papéis & Permissões",
        "emoji": "🔒",
        "desc": "Configuração de papéis e níveis de acesso por aba",
        "actions": ["read", "create", "update", "delete"],
        "route": "roles",
    },
]


# ── Helpers de Verificação de Permissões e Decoradores ───────────────────────

def user_has_permission(resource, action):
    """Verifica se o usuário logado (g.user) tem a permissão para a ação no recurso/aba.
    Considera todos os papéis atribuídos ao usuário (múltiplos papéis são permitidos).
    """
    if not g.get("user"):
        return False
    roles = usuario_roles_lista(g.user)
    if "Admin" in roles:
        return True
    if not roles:
        return False

    col_map = {"create": "can_create", "read": "can_read", "update": "can_update", "delete": "can_delete"}
    col = col_map.get(action, "can_read")

    if USE_SQLITE:
        try:
            init_db()
            conn = sqlite3.connect(DB_PATH)
            cur = conn.cursor()
            for role in roles:
                cur.execute(f"SELECT {col} FROM role_permissions WHERE role=? AND resource=?", (role, resource))
                row = cur.fetchone()
                if row and row[0] == 1:
                    conn.close()
                    return True
            conn.close()
        except Exception:
            return False
    else:
        perms_list = carregar_json("role_permissions.json", seed=[])
        for role in roles:
            entry = next((p for p in perms_list if p.get("role") == role and p.get("resource") == resource), None)
            if entry and entry.get(col) == 1:
                return True
    return False


def user_can_access_tab(tab_id):
    """Verifica se o usuário logado tem permissão para visualizar e acessar a aba no menu."""
    if not g.get("user"):
        return False
    roles = usuario_roles_lista(g.user)
    if "Admin" in roles:
        return True
    if not roles:
        return False
    if tab_id in ("adicionar", "baixa"):
        return user_has_permission(tab_id, "create") or user_has_permission(tab_id, "read")
    return user_has_permission(tab_id, "read")


def requires_roles(*allowed_roles):
    """Decorator legado para permitir acesso apenas a papéis específicos ou Admin."""
    def decorator(f):
        @wraps(f)
        def wrapped(*args, **kwargs):
            if not g.get("user"):
                return redirect(url_for("login", next=request.path))
            roles = usuario_roles_lista(g.user)
            if "Admin" in roles or any(r in allowed_roles for r in roles):
                return f(*args, **kwargs)
            flash("Acesso negado: você não tem permissão para acessar esta área.")
            return redirect(url_for("home"))
        return wrapped
    return decorator


def requires_permission(resource, action):
    """Decorador para proteger rotas baseado em permissões granulares da tabela role_permissions."""
    def decorator(f):
        @wraps(f)
        def wrapped(*args, **kwargs):
            if not g.get("user"):
                return redirect(url_for("login", next=request.path))
            if "Admin" in usuario_roles_lista(g.user):
                return f(*args, **kwargs)
            if user_has_permission(resource, action):
                return f(*args, **kwargs)
            flash("Acesso negado: você não tem permissão para acessar esta área ou realizar esta ação.")
            return redirect(url_for("home"))
        return wrapped
    return decorator


@app.context_processor
def inject_permissions():
    return dict(
        has_permission=user_has_permission,
        can_access_tab=user_can_access_tab,
        system_tabs=SYSTEM_TABS,
    )



def init_db():
    """Initialize SQLite database and tables used by the app when USE_SQLITE is enabled.
    Creates both a generic collections table (legacy) and proper normalized tables for
    materiais, produtos, pedidos, movimentacoes, sobras, despesas and usuarios.
    """
    os.makedirs(DATA_DIR, exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    # legacy generic collection storage (used for produtos, pedidos, etc.)
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS collections (
            name TEXT NOT NULL,
            id TEXT PRIMARY KEY,
            data TEXT NOT NULL
        )
        """
    )
    cur.execute("CREATE INDEX IF NOT EXISTS idx_collections_name ON collections(name)")

    # proper materiais table with explicit columns and uniqueness on (lower(nome), gtin)
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS materiais (
            id TEXT PRIMARY KEY,
            nome TEXT NOT NULL,
            categoria TEXT,
            emoji TEXT,
            quantidade REAL DEFAULT 0,
            unidade TEXT,
            quantidade_minima REAL DEFAULT 0,
            custo REAL DEFAULT 0,
            gtin TEXT,
            foto TEXT,
            created_at TEXT,
            updated_at TEXT
        )
        """
    )
    # case-insensitive uniqueness on name+gtin to avoid duplicates
    cur.execute(
        "CREATE UNIQUE INDEX IF NOT EXISTS ux_materiais_nome_gtin ON materiais (lower(nome), gtin)"
    )

    # produtos table
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS produtos (
            id TEXT PRIMARY KEY,
            nome TEXT NOT NULL,
            emoji TEXT,
            preco_venda REAL DEFAULT 0,
            receita TEXT,
            gtin TEXT,
            estoque_pronto INTEGER DEFAULT 0,
            created_at TEXT,
            updated_at TEXT
        )
        """
    )

    # pedidos table
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS pedidos (
            id TEXT PRIMARY KEY,
            cliente TEXT,
            produto_id TEXT,
            produto_nome TEXT,
            produto_emoji TEXT,
            quantidade INTEGER,
            preco_unitario REAL,
            valor_total REAL,
            status TEXT,
            materiais_baixados INTEGER DEFAULT 0,
            usou_estoque_pronto INTEGER DEFAULT 0,
            data_pedido TEXT,
            data_pedido_iso TEXT,
            observacoes TEXT,
            created_at TEXT,
            updated_at TEXT
        )
        """
    )

    # movimentacoes (audit/history)
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS movimentacoes (
            id TEXT PRIMARY KEY,
            tipo TEXT,
            material_nome TEXT,
            quantidade REAL,
            unidade TEXT,
            motivo TEXT,
            data TEXT,
            usuario TEXT,
            created_at TEXT
        )
        """
    )
    cur.execute("CREATE INDEX IF NOT EXISTS idx_movimentacoes_created ON movimentacoes(created_at)")

    # sobras
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS sobras (
            id TEXT PRIMARY KEY,
            material_id TEXT,
            descricao TEXT,
            quantidade REAL,
            unidade TEXT,
            data TEXT,
            status TEXT,
            created_at TEXT,
            updated_at TEXT
        )
        """
    )

    # despesas
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS despesas (
            id TEXT PRIMARY KEY,
            descricao TEXT,
            valor REAL,
            categoria TEXT,
            data TEXT,
            created_at TEXT
        )
        """
    )

    # usuarios
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS usuarios (
            id TEXT PRIMARY KEY,
            username TEXT UNIQUE,
            password_hash TEXT,
            role TEXT,
            nome TEXT,
            avatar TEXT,
            created_at TEXT,
            session_version INTEGER DEFAULT 0
        )
        """
    )
    # Migrations for tables
    for col, ctype in [('role', 'TEXT'), ('session_version', 'INTEGER DEFAULT 0'), ('nome', 'TEXT'), ('avatar', 'TEXT'), ('roles', 'TEXT')]:
        try:
            cur.execute(f"ALTER TABLE usuarios ADD COLUMN {col} {ctype}")
        except Exception:
            pass

    for col, ctype in [('gtin', 'TEXT'), ('estoque_pronto', 'INTEGER DEFAULT 0')]:
        try:
            cur.execute(f"ALTER TABLE produtos ADD COLUMN {col} {ctype}")
        except Exception:
            pass

    for col, ctype in [('usou_estoque_pronto', 'INTEGER DEFAULT 0')]:
        try:
            cur.execute(f"ALTER TABLE pedidos ADD COLUMN {col} {ctype}")
        except Exception:
            pass

    # roles table: defines roles and descriptions
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS roles (
            id TEXT PRIMARY KEY,
            name TEXT UNIQUE NOT NULL,
            description TEXT,
            is_system INTEGER DEFAULT 0,
            created_at TEXT,
            updated_at TEXT
        )
        """
    )

    # role_permissions table: per-role per-resource CRUD flags
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS role_permissions (
            role TEXT NOT NULL,
            resource TEXT NOT NULL,
            can_create INTEGER DEFAULT 0,
            can_read INTEGER DEFAULT 0,
            can_update INTEGER DEFAULT 0,
            can_delete INTEGER DEFAULT 0,
            updated_at TEXT,
            PRIMARY KEY (role, resource)
        )
        """
    )

    # audits table: records actor, target, action and details for sensitive changes
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS audits (
            id TEXT PRIMARY KEY,
            actor_id TEXT,
            actor_username TEXT,
            target_user_id TEXT,
            action TEXT,
            details TEXT,
            created_at TEXT
        )
        """
    )

    # relatorios_customizados table: stores user-created custom reports and chart setups
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS relatorios_customizados (
            id TEXT PRIMARY KEY,
            titulo TEXT NOT NULL,
            tipo TEXT NOT NULL,
            tipo_grafico TEXT NOT NULL,
            categoria_filtro TEXT,
            status_filtro TEXT,
            apenas_criticos INTEGER DEFAULT 0,
            observacoes TEXT,
            criado_por TEXT,
            created_at TEXT,
            updated_at TEXT
        )
        """
    )

    conn.commit()
    seed_roles_se_necessario(conn)

    # Normaliza categorias dos materiais para o conjunto canônico, mantendo tudo
    # sincronizado com as categorias existentes (ex.: "Couro" -> "Courino",
    # "Metais" -> "Metal", "Polímero" -> "Outros").
    try:
        cur.execute("SELECT id, nome, categoria FROM materiais")
        for row in cur.fetchall():
            nova = normalizar_categoria(row[2], row[1])
            if nova != row[2]:
                cur.execute("UPDATE materiais SET categoria=?, emoji=? WHERE id=?", (nova, CATEGORIAS_EMOJI.get(nova, "🔹"), row[0]))
        conn.commit()
    except Exception:
        pass

    # Migração única: limpa resíduos de dados de teste e popula itens padrão do sistema
    try:
        cur.execute("CREATE TABLE IF NOT EXISTS app_meta (key TEXT PRIMARY KEY, value TEXT)")
        cur.execute("SELECT value FROM app_meta WHERE key='seed_padrao_v5_aplicado'")
        if not cur.fetchone():
            cur.execute("DELETE FROM pedidos")
            cur.execute("DELETE FROM sobras")
            cur.execute("DELETE FROM despesas")
            cur.execute("DELETE FROM movimentacoes")
            cur.execute("DELETE FROM relatorios_customizados")
            cur.execute("DELETE FROM produtos")
            cur.execute("DELETE FROM materiais")
            cur.execute("DELETE FROM usuarios WHERE username != 'admin'")
            cur.execute("DELETE FROM roles WHERE is_system = 0")
            cur.execute("DELETE FROM role_permissions WHERE role NOT IN (SELECT name FROM roles WHERE is_system = 1)")
            cur.execute("UPDATE usuarios SET role='Admin', roles=? WHERE username='admin'", (serializar_roles(['Admin']),))

            # Se a tabela de materiais ficou vazia, carrega do SEED_FILE
            if os.path.exists(SEED_FILE):
                try:
                    with open(SEED_FILE, encoding="utf-8") as f:
                        seed_data = json.load(f)
                    now = agora().isoformat()
                    for m in seed_data:
                        _id = m.get("id") or str(uuid.uuid4())
                        cur.execute(
                            "INSERT OR IGNORE INTO materiais (id,nome,categoria,emoji,quantidade,unidade,quantidade_minima,custo,gtin,foto,created_at,updated_at) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
                            (_id, m.get("nome"), m.get("categoria"), m.get("emoji"), float(m.get("quantidade") or 0), m.get("unidade"), float(m.get("quantidade_minima") or 0), float(m.get("custo") or 0), m.get("gtin"), m.get("foto"), now, now)
                        )
                except Exception:
                    pass

            cur.execute("INSERT OR REPLACE INTO app_meta (key, value) VALUES ('seed_padrao_v5_aplicado', '1')")
            conn.commit()
    except Exception:
        pass

    conn.close()



def carregar_json(nome_arquivo, seed=None):
    """Load a list of items from JSON file or from SQLite collections table when enabled."""
    if USE_SQLITE:
        name = os.path.splitext(nome_arquivo)[0]
        init_db()
        conn = sqlite3.connect(DB_PATH)
        cur = conn.cursor()
        cur.execute("SELECT data FROM collections WHERE name=? ORDER BY rowid DESC", (name,))
        rows = cur.fetchall()
        conn.close()
        if not rows and seed is not None:
            # seed DB from provided seed value
            salvar_json(nome_arquivo, seed)
            return seed
        return [json.loads(r[0]) for r in rows]

    caminho = os.path.join(DATA_DIR, nome_arquivo)
    if not os.path.exists(caminho):
        os.makedirs(DATA_DIR, exist_ok=True)
        with open(caminho, "w", encoding="utf-8") as f:
            json.dump(seed if seed is not None else [], f, ensure_ascii=False, indent=2)
    with open(caminho, encoding="utf-8") as f:
        return json.load(f)


def salvar_json(nome_arquivo, dados):
    """Save a list of items into JSON file or into SQLite collections table when enabled.
    Each item must be a dict and will have an id field.
    """
    if USE_SQLITE:
        name = os.path.splitext(nome_arquivo)[0]
        init_db()
        conn = sqlite3.connect(DB_PATH)
        cur = conn.cursor()
        # remove previous collection rows
        cur.execute("DELETE FROM collections WHERE name=?", (name,))
        for item in dados:
            _id = item.get("id") or str(uuid.uuid4())
            item["id"] = _id
            cur.execute("INSERT INTO collections(name, id, data) VALUES (?, ?, ?)", (name, _id, json.dumps(item, ensure_ascii=False)))
        conn.commit()
        conn.close()
        return

    os.makedirs(DATA_DIR, exist_ok=True)
    caminho = os.path.join(DATA_DIR, nome_arquivo)
    with open(caminho, "w", encoding="utf-8") as f:
        json.dump(dados, f, ensure_ascii=False, indent=2)


def registrar_movimentacao(tipo, quantidade, unidade, motivo, material_nome=""):
    """Grava um evento no histórico (usado em Alertas e Relatórios). Mantém só os 200 mais recentes.
    Persiste em tabela movimentacoes quando USE_SQLITE está ativo, e mantém ainda o arquivo JSON legada
    para compatibilidade com código antigo.
    """
    usuario = session.get("user_id") if session else None
    data_text = agora().strftime("%d/%m/%Y %H:%M")

    if USE_SQLITE:
        init_db()
        conn = sqlite3.connect(DB_PATH)
        cur = conn.cursor()
        try:
            cur.execute(
                "INSERT INTO movimentacoes (id,tipo,material_nome,quantidade,unidade,motivo,data,usuario,created_at) VALUES (?,?,?,?,?,?,?,?,?)",
                (str(uuid.uuid4()), tipo, material_nome, float(quantidade or 0), unidade, motivo or "Não informado", data_text, usuario, agora().isoformat()),
            )
            # Keep only 200 latest rows
            cur.execute("DELETE FROM movimentacoes WHERE id NOT IN (SELECT id FROM movimentacoes ORDER BY created_at DESC LIMIT 200)")
            conn.commit()
        except Exception:
            conn.rollback()
        finally:
            conn.close()
        return

    # legacy JSON fallback/update (kept for templates that read movimentacoes.json)
    movs = carregar_json("movimentacoes.json")
    movs.insert(0, {
        "id": str(uuid.uuid4()),
        "tipo": tipo,
        "material_nome": material_nome,
        "quantidade": float(quantidade or 0),
        "unidade": unidade,
        "motivo": motivo or "Não informado",
        "data": data_text,
        "usuario": usuario,
    })
    salvar_json("movimentacoes.json", movs[:200])


# ── Autenticação simples (usuários em collections 'usuarios') ─────────────────

def carregar_usuarios():
    if USE_SQLITE:
        init_db()
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        cur.execute("SELECT * FROM usuarios ORDER BY username")
        rows = cur.fetchall()
        conn.close()
        return [{
            "id": r["id"],
            "username": r["username"],
            "nome": r["nome"] if ("nome" in r.keys() and r["nome"]) else "",
            "avatar": r["avatar"] if ("avatar" in r.keys() and r["avatar"]) else "",
            "password_hash": r["password_hash"],
            "role": r["role"] if r["role"] is not None else "",
            "roles": r["roles"] if ("roles" in r.keys() and r["roles"]) else "",
            "created_at": r["created_at"],
            "session_version": r["session_version"] if ("session_version" in r.keys()) else 0,
        } for r in rows]
    # legacy JSON fallback: each user dict may include role, nome, avatar
    users = carregar_json("usuarios.json", seed=[])
    for u in users:
        if "session_version" not in u:
            u["session_version"] = 0
        if "nome" not in u:
            u["nome"] = ""
        if "avatar" not in u:
            u["avatar"] = ""
        if "roles" not in u:
            u["roles"] = ""
    return users


def salvar_usuarios(usuarios):
    if USE_SQLITE:
        init_db()
        conn = sqlite3.connect(DB_PATH)
        cur = conn.cursor()
        try:
            cur.execute("BEGIN IMMEDIATE")
            cur.execute("DELETE FROM usuarios")
            now = agora().isoformat()
            for u in usuarios:
                _id = u.get("id") or str(uuid.uuid4())
                cur.execute(
                    "INSERT INTO usuarios (id, username, password_hash, role, nome, avatar, roles, created_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                    (_id, u.get("username"), u.get("password_hash"), u.get("role") or "", u.get("nome") or "", u.get("avatar") or "", serializar_roles(u.get("roles") or u.get("role") or ""), u.get("created_at") or now),
                )
            conn.commit()
        except Exception:
            conn.rollback()
            raise
        finally:
            conn.close()
        return
    return salvar_json("usuarios.json", usuarios)


def encontrar_usuario_por_username(username):
    if USE_SQLITE:
        init_db()
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        cur.execute("SELECT * FROM usuarios WHERE username=?", (username,))
        r = cur.fetchone()
        conn.close()
        if not r:
            return None
        return {
            "id": r["id"],
            "username": r["username"],
            "nome": r["nome"] if ("nome" in r.keys() and r["nome"]) else "",
            "avatar": r["avatar"] if ("avatar" in r.keys() and r["avatar"]) else "",
            "password_hash": r["password_hash"],
            "role": r["role"] if r["role"] is not None else "",
            "roles": r["roles"] if ("roles" in r.keys() and r["roles"]) else "",
            "created_at": r["created_at"],
            "session_version": r["session_version"] if ("session_version" in r.keys()) else 0,
        }

    usuarios = carregar_usuarios()
    for u in usuarios:
        if u.get("username") == username:
            return u
    return None


def seed_roles_se_necessario(conn=None):
    """Garante a existência dos papéis padrão e suas permissões canônicas."""
    close_at_end = False
    if conn is None:
        conn = sqlite3.connect(DB_PATH)
        close_at_end = True
    cur = conn.cursor()

    cur.execute("SELECT COUNT(1) FROM roles")
    count = cur.fetchone()[0]

    now = agora().isoformat()
    default_roles_data = [
        ("Admin", "Acesso irrestrito a todas as áreas, abas e configurações do sistema", 1, {
            "estoque": (1, 1, 1, 1),
            "adicionar": (1, 1, 1, 1),
            "baixa": (1, 1, 1, 1),
            "produtos": (1, 1, 1, 1),
            "pedidos": (1, 1, 1, 1),
            "sobras": (1, 1, 1, 1),
            "financeiro": (1, 1, 1, 1),
            "relatorios": (1, 1, 1, 1),
            "usuarios": (1, 1, 1, 1),
            "roles": (1, 1, 1, 1),
        }),
        ("Estoque", "Gestão de insumos, materiais, baixas manuais e sobras", 1, {
            "estoque": (1, 1, 1, 1),
            "adicionar": (1, 1, 0, 0),
            "baixa": (1, 1, 0, 0),
            "sobras": (1, 1, 1, 1),
            "produtos": (0, 1, 0, 0),
            "relatorios": (0, 1, 0, 0),
        }),
        ("Vendas", "Gestão comercial de pedidos de clientes e catálogo de produtos", 1, {
            "pedidos": (1, 1, 1, 1),
            "produtos": (0, 1, 0, 0),
            "estoque": (0, 1, 0, 0),
            "relatorios": (0, 1, 0, 0),
        }),
        ("Producao", "Acompanhamento da confecção de pedidos, receitas e baixa de insumos", 1, {
            "pedidos": (1, 1, 1, 0),
            "produtos": (1, 1, 1, 0),
            "estoque": (0, 1, 1, 0),
            "baixa": (1, 1, 0, 0),
            "sobras": (1, 1, 1, 0),
        }),
        ("Financeiro", "Fluxo financeiro, controle de despesas, faturamento e relatórios", 1, {
            "financeiro": (1, 1, 1, 1),
            "relatorios": (0, 1, 0, 0),
            "pedidos": (0, 1, 0, 0),
            "estoque": (0, 1, 0, 0),
        }),
        ("Relatorios", "Acesso analítico a alertas, relatórios gerenciais e exportações", 1, {
            "relatorios": (0, 1, 0, 0),
            "financeiro": (0, 1, 0, 0),
            "estoque": (0, 1, 0, 0),
        }),
    ]

    for name, desc, is_sys, perms in default_roles_data:
        cur.execute("SELECT id FROM roles WHERE name=?", (name,))
        existing = cur.fetchone()
        if not existing:
            role_id = str(uuid.uuid4())
            cur.execute(
                "INSERT INTO roles (id, name, description, is_system, created_at, updated_at) VALUES (?, ?, ?, ?, ?, ?)",
                (role_id, name, desc, is_sys, now, now),
            )
            for res, (c, r, u, d) in perms.items():
                cur.execute(
                    "INSERT OR REPLACE INTO role_permissions (role, resource, can_create, can_read, can_update, can_delete, updated_at) VALUES (?, ?, ?, ?, ?, ?, ?)",
                    (name, res, c, r, u, d, now),
                )
    conn.commit()

    if close_at_end:
        conn.close()


def carregar_papeis():
    """Retorna a lista de papéis com detalhes, permissões e contagem de usuários."""
    if USE_SQLITE:
        init_db()
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()

        cur.execute("SELECT * FROM roles ORDER BY CASE WHEN name='Admin' THEN 0 ELSE 1 END, name")
        roles_rows = cur.fetchall()

        cur.execute("SELECT role, COUNT(1) as cnt FROM usuarios GROUP BY role")
        user_counts = {r["role"]: r["cnt"] for r in cur.fetchall()}

        cur.execute("SELECT role, resource, can_create, can_read, can_update, can_delete FROM role_permissions")
        perms_by_role = {}
        for r in cur.fetchall():
            perms_by_role.setdefault(r["role"], {})[r["resource"]] = {
                "can_create": bool(r["can_create"]),
                "can_read": bool(r["can_read"]),
                "can_update": bool(r["can_update"]),
                "can_delete": bool(r["can_delete"]),
            }
        conn.close()

        result = []
        for r in roles_rows:
            r_name = r["name"]
            perms = perms_by_role.get(r_name, {})
            allowed_tabs = []
            for tab in SYSTEM_TABS:
                t_id = tab["id"]
                p = perms.get(t_id, {})
                if r_name == "Admin" or p.get("can_read") or (t_id in ("adicionar", "baixa") and p.get("can_create")):
                    allowed_tabs.append(tab)

            result.append({
                "id": r["id"],
                "name": r["name"],
                "description": r["description"] or "",
                "is_system": bool(r["is_system"]),
                "created_at": r["created_at"],
                "user_count": user_counts.get(r_name, 0),
                "permissions": perms,
                "allowed_tabs": allowed_tabs,
                "total_tabs": len(SYSTEM_TABS),
            })
        return result

    # Fallback JSON
    roles = carregar_json("roles.json", seed=[])
    if not roles:
        roles = [
            {"id": "admin-id", "name": "Admin", "description": "Acesso total", "is_system": True, "created_at": agora().isoformat()},
            {"id": "est-id", "name": "Estoque", "description": "Estoque e insumos", "is_system": True, "created_at": agora().isoformat()},
            {"id": "fin-id", "name": "Financeiro", "description": "Financeiro", "is_system": True, "created_at": agora().isoformat()},
        ]
        salvar_json("roles.json", roles)
    return roles


def encontrar_papel(role_name):
    """Retorna o objeto do papel e seu mapa de permissões."""
    if USE_SQLITE:
        init_db()
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        cur.execute("SELECT * FROM roles WHERE name=?", (role_name,))
        r = cur.fetchone()
        if not r:
            conn.close()
            return None
        cur.execute("SELECT resource, can_create, can_read, can_update, can_delete FROM role_permissions WHERE role=?", (role_name,))
        perms = {}
        for p in cur.fetchall():
            perms[p["resource"]] = {
                "can_create": bool(p["can_create"]),
                "can_read": bool(p["can_read"]),
                "can_update": bool(p["can_update"]),
                "can_delete": bool(p["can_delete"]),
            }
        conn.close()
        return {
            "id": r["id"],
            "name": r["name"],
            "description": r["description"] or "",
            "is_system": bool(r["is_system"]),
            "created_at": r["created_at"],
            "permissions": perms,
        }
    papeis = carregar_papeis()
    return next((p for p in papeis if p.get("name") == role_name), None)


def criar_usuario_padrao_se_necessario():
    if USE_SQLITE:
        init_db()
        conn = sqlite3.connect(DB_PATH)
        cur = conn.cursor()
        cur.execute("SELECT COUNT(1) FROM usuarios")
        cnt = cur.fetchone()[0]
        if cnt == 0:
            senha = os.environ.get("ADMIN_PASSWORD", "admin")
            uid = str(uuid.uuid4())
            cur.execute("INSERT INTO usuarios (id,username,password_hash,role,roles,created_at) VALUES (?,?,?,?,?,?)",
                        (uid, 'admin', generate_password_hash(senha), 'Admin', serializar_roles(['Admin']), agora().isoformat()))
            conn.commit()
        conn.close()
        return

    usuarios = carregar_usuarios()
    if not usuarios:
        senha = os.environ.get("ADMIN_PASSWORD", "admin")
        usuario = {
            "id": str(uuid.uuid4()),
            "username": "admin",
            "password_hash": generate_password_hash(senha),
            "role": 'Admin',
            "roles": ['Admin'],
            "created_at": agora().isoformat(),
        }
        usuarios.append(usuario)
        salvar_usuarios(usuarios)


criar_usuario_padrao_se_necessario()


def recreate_db_with_admin_forced():
    """If RECREATE_DB env var is set (1/true/yes), delete and recreate the SQLite DB and
    inject an 'admin' user with a generated password. Writes credentials to data/admin_credentials.txt
    so the operator can retrieve the password after starting the app.
    """
    if not USE_SQLITE:
        return
    if os.environ.get('RECREATE_DB', '').lower() not in ('1', 'true', 'yes'):
        return
    try:
        if os.path.exists(DB_PATH):
            os.remove(DB_PATH)
    except Exception:
        pass
    # Recreate schema
    init_db()
    pwd = secrets.token_urlsafe(10)
    uid = str(uuid.uuid4())
    now = agora().isoformat()
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    try:
        cur.execute("INSERT OR REPLACE INTO usuarios (id,username,password_hash,role,roles,created_at) VALUES (?,?,?,?,?,?)", (uid, 'admin', generate_password_hash(pwd), 'Admin', serializar_roles(['Admin']), now))
        conn.commit()
    finally:
        conn.close()
    os.makedirs(DATA_DIR, exist_ok=True)
    cred_path = os.path.join(DATA_DIR, 'admin_credentials.txt')
    try:
        with open(cred_path, 'w', encoding='utf-8') as f:
            f.write(f"username: admin\npassword: {pwd}\ncreated_at: {now}\n")
    except Exception:
        pass
    # Also print so logs/console show it when app starts
    print(f"RECREATE_DB: created {DB_PATH} and wrote admin credentials to {cred_path}")

# If the deploy/start command explicitly asks for recreation, do it now
if os.environ.get('RECREATE_DB', '').lower() in ('1','true','yes'):
    recreate_db_with_admin_forced()


@app.before_request
def require_login():
    # Allow these endpoints unauthenticated
    allowed = {"login", "static", "em_construcao", "uploaded_file"}
    if request.endpoint is None:
        return
    if request.endpoint in allowed:
        return
    # load user into g for template access and role checks
    if session.get("user_id"):
        user_id = session.get("user_id")
        # try to find user by id
        user = None
        if USE_SQLITE:
            init_db()
            conn = sqlite3.connect(DB_PATH)
            conn.row_factory = sqlite3.Row
            cur = conn.cursor()
            cur.execute("SELECT * FROM usuarios WHERE id=?", (user_id,))
            r = cur.fetchone()
            conn.close()
            if r:
                user = {
                    "id": r["id"],
                    "username": r["username"],
                    "nome": r["nome"] if ("nome" in r.keys() and r["nome"]) else "",
                    "avatar": r["avatar"] if ("avatar" in r.keys() and r["avatar"]) else "",
                    "role": r["role"] if r["role"] is not None else "",
                    "roles": r["roles"] if ("roles" in r.keys() and r["roles"]) else "",
                    "session_version": r["session_version"] if ("session_version" in r.keys()) else 0,
                }
        else:
            usuarios = carregar_usuarios()
            user = next((u for u in usuarios if u.get("id") == user_id), None)
        # session invalidation: compare session_version stored in session with DB; if mismatch, force logout
        db_ver = (user.get('session_version') if user else 0)
        sess_ver = session.get('session_version')
        if sess_ver is None or sess_ver != db_ver:
            # expire session
            session.pop('user_id', None)
            session.pop('session_version', None)
            flash('Sessão expirada. Por favor faça login novamente.')
            return redirect(url_for('login'))
        g.user = user
        return
    return redirect(url_for("login", next=request.path))


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        senha = request.form.get("password", "")
        user = encontrar_usuario_por_username(username)
        if user:
            stored = user.get("password_hash", "")
            # Backward/compatibility: allow a stored plaintext marker 'PLAIN:' for quick local setup
            if stored.startswith('PLAIN:'):
                if stored[len('PLAIN:'):] == senha:
                    session["user_id"] = user["id"]
                    # persist session version to allow server-side invalidation
                    session['session_version'] = user.get('session_version', 0) if isinstance(user, dict) else 0
                    flash("Autenticado com sucesso.")
                    nxt = request.args.get("next") or url_for("home")
                    return redirect(nxt)
            else:
                if check_password_hash(stored, senha):
                    session["user_id"] = user["id"]
                    session['session_version'] = user.get('session_version', 0) if isinstance(user, dict) else 0
                    flash("Autenticado com sucesso.")
                    nxt = request.args.get("next") or url_for("home")
                    return redirect(nxt)
        flash("Usuário ou senha inválidos.")
        return redirect(url_for("login"))
    return render_template("login.html")


@app.route("/logout")
def logout():
    session.pop("user_id", None)
    flash("Desconectado.")
    return redirect(url_for("login"))


# Minha conta (Editar Perfil, Nome, Foto e Trocar Senha)
@app.route('/minha-conta', methods=['GET', 'POST'])
def minha_conta():
    if not g.get('user'):
        return redirect(url_for('login'))

    user_id = g.user.get('id')
    user = encontrar_usuario_por_username(g.user.get('username'))
    if not user:
        flash('Usuário não encontrado.')
        return redirect(url_for('login'))

    papeis = carregar_papeis()
    papeis_map = {p["name"]: p for p in papeis}
    user_role_info = papeis_map.get(user.get("role"))

    if request.method == 'POST':
        nome = request.form.get('nome', '').strip()
        remover_avatar = (request.form.get('remover_avatar') == '1')
        current_pwd = request.form.get('current_password', '')
        new_pwd = request.form.get('new_password', '')
        new_pwd2 = request.form.get('new_password2', '')

        # 1. Tratar alteração de senha (se algum campo de nova senha for preenchido)
        password_changed = False
        new_hash = None
        if new_pwd or new_pwd2 or current_pwd:
            if not current_pwd:
                flash('Para alterar a senha, informe sua senha atual.')
                return render_template('minha_conta.html', user=user, user_role_info=user_role_info, system_tabs=SYSTEM_TABS)
            if new_pwd != new_pwd2:
                flash('A nova senha e a confirmação não conferem.')
                return render_template('minha_conta.html', user=user, user_role_info=user_role_info, system_tabs=SYSTEM_TABS)
            if not new_pwd:
                flash('A nova senha não pode estar em branco.')
                return render_template('minha_conta.html', user=user, user_role_info=user_role_info, system_tabs=SYSTEM_TABS)

            stored = user.get('password_hash', '')
            ok = False
            if stored.startswith('PLAIN:'):
                ok = (stored[len('PLAIN:'):] == current_pwd)
            else:
                ok = check_password_hash(stored, current_pwd)
            if not ok:
                flash('Senha atual incorreta.')
                return render_template('minha_conta.html', user=user, user_role_info=user_role_info, system_tabs=SYSTEM_TABS)

            new_hash = generate_password_hash(new_pwd)
            password_changed = True

        # 2. Tratar Foto de Perfil (Avatar)
        current_avatar = user.get('avatar') or ''
        new_avatar = current_avatar

        if remover_avatar:
            if current_avatar:
                old_path = os.path.join(DATA_DIR, 'uploads', current_avatar)
                try:
                    if os.path.exists(old_path):
                        os.remove(old_path)
                except Exception:
                    pass
            new_avatar = ''
        elif 'avatar' in request.files:
            f = request.files.get('avatar')
            if f and f.filename:
                os.makedirs(os.path.join(DATA_DIR, 'uploads'), exist_ok=True)
                if current_avatar:
                    old_path = os.path.join(DATA_DIR, 'uploads', current_avatar)
                    try:
                        if os.path.exists(old_path):
                            os.remove(old_path)
                    except Exception:
                        pass
                ext = os.path.splitext(secure_filename(f.filename))[1].lower()
                clean_name = f"avatar_{user_id}_{int(agora().timestamp())}{ext}"
                target_path = os.path.join(DATA_DIR, 'uploads', clean_name)
                f.save(target_path)
                new_avatar = clean_name

        # 3. Salvar no Banco
        if USE_SQLITE:
            init_db()
            conn = sqlite3.connect(DB_PATH)
            cur = conn.cursor()
            try:
                if password_changed:
                    cur.execute("UPDATE usuarios SET nome=?, avatar=?, password_hash=? WHERE id=?",
                                (nome, new_avatar, new_hash, user_id))
                else:
                    cur.execute("UPDATE usuarios SET nome=?, avatar=? WHERE id=?",
                                (nome, new_avatar, user_id))
                conn.commit()
                try:
                    details = f"nome={nome};avatar={'yes' if new_avatar else 'no'};password_changed={'yes' if password_changed else 'no'}"
                    cur.execute("INSERT INTO audits (id, actor_id, actor_username, target_user_id, action, details, created_at) VALUES (?,?,?,?,?,?,?)",
                                (str(uuid.uuid4()), user_id, user.get('username'), user_id, 'update_profile', details, agora().isoformat()))
                    conn.commit()
                except Exception:
                    pass
            except Exception as e:
                conn.rollback()
                flash('Erro ao salvar perfil: ' + str(e))
                return render_template('minha_conta.html', user=user, user_role_info=user_role_info, system_tabs=SYSTEM_TABS)
            finally:
                conn.close()
        else:
            usuarios = carregar_usuarios()
            for u in usuarios:
                if u.get('id') == user_id:
                    u['nome'] = nome
                    u['avatar'] = new_avatar
                    if password_changed:
                        u['password_hash'] = new_hash
            salvar_usuarios(usuarios)

        g.user['nome'] = nome
        g.user['avatar'] = new_avatar

        if password_changed:
            flash('Perfil e senha atualizados com sucesso!')
        else:
            flash('Perfil atualizado com sucesso!')
        return redirect(url_for('minha_conta'))

    return render_template('minha_conta.html', user=user, user_role_info=user_role_info, system_tabs=SYSTEM_TABS)


# ── Usuários (Admin / Gestão de Contas) ───────────────────────────────────────
@app.route("/usuarios")
@requires_permission("usuarios", "read")
def usuarios():
    usuarios_lista = carregar_usuarios()
    papeis = carregar_papeis()
    papeis_map = {p["name"]: p for p in papeis}
    return render_template("usuarios.html", usuarios=usuarios_lista, papeis_map=papeis_map)


@app.route("/usuarios/novo", methods=["GET", "POST"])
@requires_permission("usuarios", "create")
def usuarios_novo():
    papeis = carregar_papeis()

    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        password2 = request.form.get("password2", "")
        roles = [r for r in request.form.getlist("roles") if r.strip()]
        if not roles:
            legado = request.form.get("role", "").strip()
            if legado:
                roles = [legado]
        role = roles[0] if roles else ""

        if not username or not password:
            flash("Nome de usuário e senha são obrigatórios.")
            return render_template("usuario_form.html", papeis=papeis)
        if password != password2:
            flash("As senhas não conferem.")
            return render_template("usuario_form.html", papeis=papeis)
        exists = encontrar_usuario_por_username(username)
        if exists:
            flash("Já existe um usuário com este nome.")
            return render_template("usuario_form.html", papeis=papeis)

        now = agora().isoformat()
        uid = str(uuid.uuid4())
        password_hash = generate_password_hash(password)

        if USE_SQLITE:
            init_db()
            conn = sqlite3.connect(DB_PATH)
            cur = conn.cursor()
            try:
                cur.execute(
                    "INSERT INTO usuarios (id, username, password_hash, role, roles, created_at) VALUES (?, ?, ?, ?, ?, ?)",
                    (uid, username, password_hash, role, serializar_roles(roles), now),
                )
                conn.commit()
                try:
                    cur.execute(
                        "INSERT INTO audits (id, actor_id, actor_username, target_user_id, action, details, created_at) VALUES (?, ?, ?, ?, ?, ?, ?)",
                        (str(uuid.uuid4()), session.get("user_id"), g.user.get("username") if g.get("user") else None, uid, "create_user", f"roles={roles}", agora().isoformat()),
                    )
                    conn.commit()
                except Exception:
                    pass
            except Exception as e:
                conn.rollback()
                flash("Erro ao criar usuário: " + str(e))
                return render_template("usuario_form.html", papeis=papeis)
            finally:
                conn.close()
        else:
            usuarios_lista = carregar_usuarios()
            usuarios_lista.append({
                "id": uid,
                "username": username,
                "password_hash": password_hash,
                "role": role,
                "roles": roles,
                "created_at": now,
            })
            salvar_usuarios(usuarios_lista)

        flash("Usuário criado com sucesso.")
        return redirect(url_for("usuarios"))

    return render_template("usuario_form.html", papeis=papeis)


@app.route("/usuarios/<user_id>/editar", methods=["GET", "POST"])
@requires_permission("usuarios", "update")
def usuarios_editar(user_id):
    papeis = carregar_papeis()
    if USE_SQLITE:
        init_db()
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        cur.execute("SELECT * FROM usuarios WHERE id=?", (user_id,))
        r = cur.fetchone()
        conn.close()
        if not r:
            flash("Usuário não encontrado.")
            return redirect(url_for("usuarios"))
        usuario = {"id": r["id"], "username": r["username"], "role": r["role"] or "", "roles": r["roles"] if ("roles" in r.keys() and r["roles"]) else ""}
    else:
        usuarios_lista = carregar_usuarios()
        usuario = next((u for u in usuarios_lista if u.get("id") == user_id), None)
        if not usuario:
            flash("Usuário não encontrado.")
            return redirect(url_for("usuarios"))

    if request.method == "POST":
        roles = [x for x in request.form.getlist("roles") if x.strip()]
        if not roles:
            legado = request.form.get("role", "").strip()
            if legado:
                roles = [legado]
        role = roles[0] if roles else ""
        new_pwd = request.form.get("password", "")
        old_roles = usuario_roles_lista(usuario)
        roles_changed = (set(roles) != set(old_roles))

        if USE_SQLITE:
            conn = sqlite3.connect(DB_PATH)
            cur = conn.cursor()
            try:
                if new_pwd:
                    if roles_changed:
                        cur.execute(
                            "UPDATE usuarios SET role=?, roles=?, password_hash=?, session_version=COALESCE(session_version,0)+1 WHERE id=?",
                            (role, serializar_roles(roles), generate_password_hash(new_pwd), user_id),
                        )
                    else:
                        cur.execute(
                            "UPDATE usuarios SET role=?, roles=?, password_hash=? WHERE id=?",
                            (role, serializar_roles(roles), generate_password_hash(new_pwd), user_id),
                        )
                else:
                    if roles_changed:
                        cur.execute(
                            "UPDATE usuarios SET role=?, roles=?, session_version=COALESCE(session_version,0)+1 WHERE id=?",
                            (role, serializar_roles(roles), user_id),
                        )
                    else:
                        cur.execute("UPDATE usuarios SET role=?, roles=? WHERE id=?", (role, serializar_roles(roles), user_id))
                conn.commit()
                try:
                    details = f"roles={roles};password_changed={'yes' if new_pwd else 'no'}"
                    cur.execute(
                        "INSERT INTO audits (id, actor_id, actor_username, target_user_id, action, details, created_at) VALUES (?, ?, ?, ?, ?, ?, ?)",
                        (str(uuid.uuid4()), session.get("user_id"), g.user.get("username") if g.get("user") else None, user_id, "update_user", details, agora().isoformat()),
                    )
                    conn.commit()
                except Exception:
                    pass
            except Exception as e:
                conn.rollback()
                flash("Erro ao atualizar usuário: " + str(e))
                return render_template("usuario_edit.html", usuario=usuario, papeis=papeis, usuario_roles=usuario_roles_lista(usuario))
            finally:
                conn.close()
        else:
            usuarios_lista = carregar_usuarios()
            for u in usuarios_lista:
                if u.get("id") == user_id:
                    u["role"] = role
                    u["roles"] = roles
                    if new_pwd:
                        u["password_hash"] = generate_password_hash(new_pwd)
            salvar_usuarios(usuarios_lista)

        flash("Usuário atualizado com sucesso.")
        return redirect(url_for("usuarios"))

    return render_template("usuario_edit.html", usuario=usuario, papeis=papeis, usuario_roles=usuario_roles_lista(usuario))


@app.route("/usuarios/<user_id>/excluir", methods=["POST"])
@requires_permission("usuarios", "delete")
def usuarios_excluir(user_id):
    usuarios_lista = carregar_usuarios()
    if len(usuarios_lista) <= 1:
        flash("Não é possível remover o último usuário do sistema.")
        return redirect(url_for("usuarios"))

    if g.get("user") and g.user.get("id") == user_id:
        flash("Você não pode excluir sua própria conta.")
        return redirect(url_for("usuarios"))

    if USE_SQLITE:
        init_db()
        conn = sqlite3.connect(DB_PATH)
        cur = conn.cursor()
        try:
            cur.execute("DELETE FROM usuarios WHERE id=?", (user_id,))
            conn.commit()
            try:
                cur.execute(
                    "INSERT INTO audits (id, actor_id, actor_username, target_user_id, action, details, created_at) VALUES (?, ?, ?, ?, ?, ?, ?)",
                    (str(uuid.uuid4()), session.get("user_id"), g.user.get("username") if g.get("user") else None, user_id, "delete_user", "", agora().isoformat()),
                )
                conn.commit()
            except Exception:
                pass
        except Exception as e:
            conn.rollback()
            flash("Erro ao remover usuário: " + str(e))
            return redirect(url_for("usuarios"))
        finally:
            conn.close()
    else:
        usuarios_lista = [u for u in usuarios_lista if u.get("id") != user_id]
        salvar_usuarios(usuarios_lista)

    flash("Usuário removido com sucesso.")
    return redirect(url_for("usuarios"))


# ── Gestão de Papéis e Permissões por Aba ─────────────────────────────────────
@app.route("/roles")
@requires_permission("roles", "read")
def roles():
    papeis = carregar_papeis()
    return render_template("roles.html", papeis=papeis, system_tabs=SYSTEM_TABS)


@app.route("/roles/novo", methods=["GET", "POST"])
@requires_permission("roles", "create")
def roles_novo():
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        description = request.form.get("description", "").strip()

        if not name:
            flash("O nome do papel é obrigatório.")
            return render_template("role_form.html", role=None, system_tabs=SYSTEM_TABS, is_novo=True)

        existing = encontrar_papel(name)
        if existing:
            flash(f'Já existe um papel cadastrado com o nome "{name}".')
            return render_template("role_form.html", role=None, system_tabs=SYSTEM_TABS, is_novo=True)

        now = agora().isoformat()
        role_id = str(uuid.uuid4())

        if USE_SQLITE:
            init_db()
            conn = sqlite3.connect(DB_PATH)
            cur = conn.cursor()
            try:
                cur.execute(
                    "INSERT INTO roles (id, name, description, is_system, created_at, updated_at) VALUES (?, ?, ?, 0, ?, ?)",
                    (role_id, name, description, now, now),
                )
                for tab in SYSTEM_TABS:
                    t_id = tab["id"]
                    can_create = 1 if request.form.get(f"can_create_{t_id}") else 0
                    can_read = 1 if request.form.get(f"can_read_{t_id}") else 0
                    can_update = 1 if request.form.get(f"can_update_{t_id}") else 0
                    can_delete = 1 if request.form.get(f"can_delete_{t_id}") else 0

                    if can_create or can_read or can_update or can_delete:
                        cur.execute(
                            "INSERT OR REPLACE INTO role_permissions (role, resource, can_create, can_read, can_update, can_delete, updated_at) VALUES (?, ?, ?, ?, ?, ?, ?)",
                            (name, t_id, can_create, can_read, can_update, can_delete, now),
                        )
                conn.commit()
                try:
                    cur.execute(
                        "INSERT INTO audits (id, actor_id, actor_username, target_user_id, action, details, created_at) VALUES (?, ?, ?, ?, ?, ?, ?)",
                        (str(uuid.uuid4()), session.get("user_id"), g.user.get("username") if g.get("user") else None, None, "create_role", f"role={name}", now),
                    )
                    conn.commit()
                except Exception:
                    pass
            except Exception as e:
                conn.rollback()
                flash("Erro ao salvar papel: " + str(e))
                return render_template("role_form.html", role=None, system_tabs=SYSTEM_TABS, is_novo=True)
            finally:
                conn.close()

        flash(f'Papel "{name}" criado com sucesso.')
        return redirect(url_for("roles"))

    return render_template("role_form.html", role=None, system_tabs=SYSTEM_TABS, is_novo=True)


@app.route("/roles/<role>/editar", methods=["GET", "POST"])
@app.route("/roles/<path:role>/editar", methods=["GET", "POST"])
@requires_permission("roles", "update")
def roles_editar(role):
    role_obj = encontrar_papel(role)
    if not role_obj:
        flash("Papel não encontrado.")
        return redirect(url_for("roles"))

    if request.method == "POST":
        new_name = request.form.get("name", "").strip() or role
        description = request.form.get("description", "").strip()

        now = agora().isoformat()
        if USE_SQLITE:
            init_db()
            conn = sqlite3.connect(DB_PATH)
            cur = conn.cursor()
            try:
                if new_name != role and role != "Admin":
                    cur.execute("SELECT id FROM roles WHERE name=? AND name!=?", (new_name, role))
                    if cur.fetchone():
                        conn.close()
                        flash(f'Já existe outro papel com o nome "{new_name}".')
                        return render_template("role_form.html", role=role_obj, system_tabs=SYSTEM_TABS, is_novo=False)
                    cur.execute("UPDATE roles SET name=?, description=?, updated_at=? WHERE name=?", (new_name, description, now, role))
                    cur.execute("UPDATE usuarios SET role=? WHERE role=?", (new_name, role))
                    cur.execute("UPDATE role_permissions SET role=? WHERE role=?", (new_name, role))
                else:
                    cur.execute("UPDATE roles SET description=?, updated_at=? WHERE name=?", (description, now, role))

                target_role = new_name if (new_name != role and role != "Admin") else role

                if target_role == "Admin":
                    for tab in SYSTEM_TABS:
                        cur.execute(
                            "INSERT OR REPLACE INTO role_permissions (role, resource, can_create, can_read, can_update, can_delete, updated_at) VALUES (?, ?, 1, 1, 1, 1, ?)",
                            ("Admin", tab["id"], now),
                        )
                else:
                    cur.execute("DELETE FROM role_permissions WHERE role=?", (target_role,))
                    for tab in SYSTEM_TABS:
                        t_id = tab["id"]
                        can_create = 1 if request.form.get(f"can_create_{t_id}") else 0
                        can_read = 1 if request.form.get(f"can_read_{t_id}") else 0
                        can_update = 1 if request.form.get(f"can_update_{t_id}") else 0
                        can_delete = 1 if request.form.get(f"can_delete_{t_id}") else 0

                        if can_create or can_read or can_update or can_delete:
                            cur.execute(
                                "INSERT INTO role_permissions (role, resource, can_create, can_read, can_update, can_delete, updated_at) VALUES (?, ?, ?, ?, ?, ?, ?)",
                                (target_role, t_id, can_create, can_read, can_update, can_delete, now),
                            )

                cur.execute("UPDATE usuarios SET session_version=COALESCE(session_version,0)+1 WHERE role=?", (target_role,))
                conn.commit()

                try:
                    cur.execute(
                        "INSERT INTO audits (id, actor_id, actor_username, target_user_id, action, details, created_at) VALUES (?, ?, ?, ?, ?, ?, ?)",
                        (str(uuid.uuid4()), session.get("user_id"), g.user.get("username") if g.get("user") else None, None, "update_role_permissions", f"role={target_role}", now),
                    )
                    conn.commit()
                except Exception:
                    pass
            except Exception as e:
                conn.rollback()
                flash("Erro ao salvar permissões do papel: " + str(e))
                return render_template("role_form.html", role=role_obj, system_tabs=SYSTEM_TABS, is_novo=False)
            finally:
                conn.close()

        flash(f'Permissões do papel "{role}" atualizadas.')
        return redirect(url_for("roles"))

    return render_template("role_form.html", role=role_obj, system_tabs=SYSTEM_TABS, is_novo=False)


@app.route("/roles/<role>/excluir", methods=["POST"])
@app.route("/roles/<path:role>/excluir", methods=["POST"])
@requires_permission("roles", "delete")
def roles_excluir(role):
    if role == "Admin":
        flash("O papel Administrador é protegido pelo sistema e não pode ser excluído.")
        return redirect(url_for("roles"))

    if USE_SQLITE:
        init_db()
        conn = sqlite3.connect(DB_PATH)
        cur = conn.cursor()
        try:
            cur.execute("UPDATE usuarios SET role='', session_version=COALESCE(session_version,0)+1 WHERE role=?", (role,))
            cur.execute("DELETE FROM role_permissions WHERE role=?", (role,))
            cur.execute("DELETE FROM roles WHERE name=?", (role,))
            conn.commit()
            try:
                cur.execute(
                    "INSERT INTO audits (id, actor_id, actor_username, target_user_id, action, details, created_at) VALUES (?, ?, ?, ?, ?, ?, ?)",
                    (str(uuid.uuid4()), session.get("user_id"), g.user.get("username") if g.get("user") else None, None, "delete_role", f"role={role}", agora().isoformat()),
                )
                conn.commit()
            except Exception:
                pass
        except Exception as e:
            conn.rollback()
            flash("Erro ao excluir papel: " + str(e))
            return redirect(url_for("roles"))
        finally:
            conn.close()

    flash(f'Papel "{role}" removido com sucesso.')
    return redirect(url_for("roles"))


@app.route("/roles/assign", methods=["GET", "POST"])
@requires_permission("roles", "update")
def roles_assign():
    """Atribuição em massa de papéis a usuários."""
    papeis = carregar_papeis()
    usuarios_lista = carregar_usuarios()

    if request.method == "POST":
        if USE_SQLITE:
            init_db()
            conn = sqlite3.connect(DB_PATH)
            cur = conn.cursor()
            try:
                changed = 0
                for u in usuarios_lista:
                    new_roles = [x for x in request.form.getlist("role_" + u["id"]) if x.strip()]
                    old_roles = usuario_roles_lista(u)
                    if set(new_roles) != set(old_roles):
                        role_principal = new_roles[0] if new_roles else ""
                        cur.execute(
                            "UPDATE usuarios SET role=?, roles=?, session_version=COALESCE(session_version,0)+1 WHERE id=?",
                            (role_principal, serializar_roles(new_roles), u["id"]),
                        )
                        changed += 1
                        try:
                            cur.execute(
                                "INSERT INTO audits (id, actor_id, actor_username, target_user_id, action, details, created_at) VALUES (?, ?, ?, ?, ?, ?, ?)",
                                (str(uuid.uuid4()), session.get("user_id"), g.user.get("username") if g.get("user") else None, u["id"], "assign_role", f"roles={new_roles}", agora().isoformat()),
                            )
                        except Exception:
                            pass
                conn.commit()
                if changed:
                    flash("Atribuições de papéis atualizadas com sucesso.")
                else:
                    flash("Nenhuma alteração nas atribuições.")
            except Exception as e:
                conn.rollback()
                flash("Erro ao atualizar atribuições: " + str(e))
            finally:
                conn.close()
        return redirect(url_for("roles"))

    return render_template("roles_assign.html", users=usuarios_lista, papeis=papeis)



# ── Materiais (estoque) ───────────────────────────────────────────────────────
def carregar_materiais():
    # When using SQLite, read from the proper 'materiais' table with a transaction.
    if USE_SQLITE:
        init_db()
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        cur.execute("SELECT COUNT(1) as cnt FROM materiais")
        cnt = cur.fetchone()[0]
        if cnt == 0 and os.path.exists(SEED_FILE):
            # seed the table from the JSON seed file
            try:
                with open(SEED_FILE, encoding="utf-8") as f:
                    seed = json.load(f)
            except Exception:
                seed = []
            if seed:
                now = agora().isoformat()
                to_insert = []
                for m in seed:
                    _id = m.get("id") or str(uuid.uuid4())
                    to_insert.append((
                        _id,
                        m.get("nome"),
                        m.get("categoria"),
                        m.get("emoji"),
                        float(m.get("quantidade") or 0),
                        m.get("unidade"),
                        float(m.get("quantidade_minima") or 0),
                        float(m.get("custo") or 0),
                        m.get("gtin"),
                        m.get("foto"),
                        now,
                        now,
                    ))
                cur.executemany(
                    "INSERT OR IGNORE INTO materiais (id,nome,categoria,emoji,quantidade,unidade,quantidade_minima,custo,gtin,foto,created_at,updated_at) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
                    to_insert,
                )
                conn.commit()
        cur.execute("SELECT * FROM materiais ORDER BY nome COLLATE NOCASE")
        rows = cur.fetchall()
        conn.close()
        # convert rows to dicts matching previous JSON structure
        result = []
        for r in rows:
            result.append({
                "id": r["id"],
                "nome": r["nome"],
                "categoria": r["categoria"],
                "emoji": r["emoji"],
                "quantidade": r["quantidade"],
                "unidade": r["unidade"],
                "quantidade_minima": r["quantidade_minima"],
                "custo": r["custo"],
                "gtin": r["gtin"],
                "foto": r["foto"],
            })
        return result

    if not os.path.exists(DATA_FILE):
        os.makedirs(DATA_DIR, exist_ok=True)
        with open(SEED_FILE, encoding="utf-8") as f:
            seed = json.load(f)
        with open(DATA_FILE, "w", encoding="utf-8") as f:
            json.dump(seed, f, ensure_ascii=False, indent=2)
    with open(DATA_FILE, encoding="utf-8") as f:
        return json.load(f)


def salvar_materiais(materiais):
    if USE_SQLITE:
        init_db()
        conn = sqlite3.connect(DB_PATH)
        cur = conn.cursor()
        try:
            # Use a transaction to atomically replace the materials set.
            cur.execute("BEGIN IMMEDIATE")
            # We'll upsert per id to preserve uniqueness constraints
            # Clear names that are no longer present
            incoming_ids = [m.get("id") for m in materiais if m.get("id")]
            if incoming_ids:
                # delete any rows not in incoming_ids
                placeholders = ",".join(["?" for _ in incoming_ids])
                cur.execute(f"DELETE FROM materiais WHERE id NOT IN ({placeholders})", incoming_ids)
            else:
                cur.execute("DELETE FROM materiais")

            now = agora().isoformat()
            for m in materiais:
                _id = m.get("id") or str(uuid.uuid4())
                cur.execute(
                    "INSERT OR REPLACE INTO materiais (id,nome,categoria,emoji,quantidade,unidade,quantidade_minima,custo,gtin,foto,created_at,updated_at) VALUES (?,?,?,?,?,?,?,?,?,?,COALESCE((SELECT created_at FROM materiais WHERE id=?),?),?)",
                    (
                        _id,
                        m.get("nome"),
                        m.get("categoria"),
                        m.get("emoji"),
                        float(m.get("quantidade") or 0),
                        m.get("unidade"),
                        float(m.get("quantidade_minima") or 0),
                        float(m.get("custo") or 0),
                        m.get("gtin"),
                        m.get("foto"),
                        _id,
                        now,
                        now,
                    ),
                )
            conn.commit()
        except Exception:
            conn.rollback()
            raise
        finally:
            conn.close()
        return

    os.makedirs(DATA_DIR, exist_ok=True)
    with open(DATA_FILE, "w", encoding="utf-8") as f:
        json.dump(materiais, f, ensure_ascii=False, indent=2)


def encontrar(materiais, material_id):
    for m in materiais:
        if m["id"] == material_id:
            return m
    return None


@app.route("/")
def home():
    materiais = carregar_materiais()
    total_itens = len(materiais)
    baixo_estoque = [m for m in materiais if m["quantidade"] <= m["quantidade_minima"]]
    return render_template("home.html", total_itens=total_itens, baixo_estoque=baixo_estoque)


@app.route("/estoque")
@requires_permission('estoque', 'read')
def estoque():
    materiais = carregar_materiais()
    cat = request.args.get("cat", "Todos")
    q = request.args.get("q", "").strip().lower()

    resultado = materiais
    if cat != "Todos":
        resultado = [m for m in resultado if m.get("categoria") == cat]
    if q:
        # busca por nome OU por código GTIN
        resultado = [
            m for m in resultado
            if q in m.get("nome", "").lower() or q in (m.get("gtin") or "").lower()
        ]

    # Categorias sincronizadas e dinâmicas com os materiais cadastrados
    categorias_disponiveis = list(dict.fromkeys(CATEGORIAS + [m.get("categoria") for m in materiais if m.get("categoria")]))

    return render_template(
        "estoque.html",
        materiais=resultado,
        categorias=categorias_disponiveis,
        cat_ativa=cat,
        q=request.args.get("q", ""),
        total=len(materiais),
    )


@app.route("/estoque/<material_id>/entrada", methods=["POST"])
@requires_permission('estoque', 'update')
def estoque_entrada(material_id):
    materiais = carregar_materiais()
    m = encontrar(materiais, material_id)
    if m:
        qtd = parse_float_ptbr(request.form.get("quantidade", 0))
        motivo = request.form.get("motivo", "").strip()
        if qtd <= 0:
            flash("Informe uma quantidade válida e maior que zero para a entrada.")
            return redirect(url_for("estoque"))

        unidade = (m.get("unidade") or "").lower()
        is_inteiro = unidade.startswith("unid") or unidade in ("unidades", "pares", "pacotes", "unidade")

        if is_inteiro:
            if abs(qtd - round(qtd)) > 1e-6:
                flash(f"Para o material '{m.get('nome')}' (unidade: {m.get('unidade')}), utilize apenas números inteiros.")
                return redirect(url_for("estoque"))
            qtd = int(round(qtd))
            m["quantidade"] = int(m.get("quantidade") or 0) + qtd
        else:
            m["quantidade"] = round(float(m.get("quantidade") or 0) + qtd, 3)

        salvar_materiais(materiais)
        registrar_movimentacao("entrada", qtd, m["unidade"], motivo or "Entrada manual de estoque", m["nome"])
        flash(f"Entrada de {qtd} {m['unidade']} registrada em {m['nome']}.")
    return redirect(url_for("estoque"))


@app.route("/estoque/<material_id>/excluir", methods=["POST"])
@requires_permission('estoque', 'delete')
def estoque_excluir(material_id):
    # Remove material and its photo (if present). Works with JSON or SQLite backend.
    if USE_SQLITE:
        init_db()
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        cur.execute("SELECT foto FROM materiais WHERE id=?", (material_id,))
        r = cur.fetchone()
        if r and r["foto"]:
            caminho = os.path.join(DATA_DIR, 'uploads', r["foto"])
            try:
                if os.path.exists(caminho):
                    os.remove(caminho)
            except Exception:
                pass
        cur.execute("DELETE FROM materiais WHERE id=?", (material_id,))
        conn.commit()
        conn.close()
        flash("Material removido.")
        return redirect(url_for("estoque"))

    materiais = carregar_materiais()
    to_remove = next((m for m in materiais if m["id"] == material_id), None)
    if to_remove and to_remove.get("foto"):
        caminho = os.path.join(DATA_DIR, 'uploads', to_remove.get('foto'))
        try:
            if os.path.exists(caminho):
                os.remove(caminho)
        except Exception:
            pass
    materiais = [m for m in materiais if m["id"] != material_id]
    salvar_materiais(materiais)
    flash("Material removido.")
    return redirect(url_for("estoque"))


@app.route("/adicionar", methods=["GET", "POST"])
@requires_permission('adicionar', 'create')
def adicionar():
    materiais = carregar_materiais()

    if request.method == "POST":
        # Basic form token to avoid duplicate submissions
        token = request.form.get('form_token')
        expected = session.pop('form_token', None)
        if not token or token != expected:
            flash('Formulário já foi enviado ou token inválido. Por favor, tente novamente.')
            return redirect(url_for('adicionar'))

        nome = request.form.get("nome", "").strip()
        categoria = request.form.get("categoria", "Outros")
        # Se selecionou "Outros" e digitou uma categoria customizada, usar ela
        if categoria == "Outros":
            cat_custom = request.form.get("categoria_custom", "").strip()
            if cat_custom:
                categoria = cat_custom
            else:
                # Sugere automaticamente a categoria a partir do nome do material,
                # mantendo o estoque sempre consistente (courino -> Courino, etc.)
                categoria = sugerir_categoria(nome)
        # Normaliza a categoria escolhida para o conjunto canônico
        categoria = normalizar_categoria(categoria, nome)
        gtin = request.form.get("gtin", "").strip()

        quantidade = parse_float_ptbr(request.form.get("quantidade", 0))
        quantidade_minima = parse_float_ptbr(request.form.get("quantidade_minima", 5))
        custo = parse_float_ptbr(request.form.get("custo", 0))

        # Basic duplicate prevention: same name + gtin
        existe = next((m for m in materiais if m["nome"].strip().lower() == nome.lower() and (m.get("gtin") or "") == gtin), None)
        if existe:
            flash("Material com mesmo nome/GTIN já existe no estoque.")
            return redirect(url_for("adicionar"))

        novo_id = str(uuid.uuid4())
        novo = {
            "id": novo_id,
            "nome": nome,
            "categoria": categoria,
            "emoji": CATEGORIAS_EMOJI.get(categoria, "🔹"),
            "quantidade": quantidade,
            "unidade": request.form.get("unidade", "unidades"),
            "quantidade_minima": quantidade_minima,
            "custo": custo,
            "gtin": gtin,
        }

        # Handle optional photo upload
        foto = None
        if 'foto' in request.files:
            f = request.files.get('foto')
            if f and f.filename:
                os.makedirs(os.path.join(DATA_DIR, 'uploads'), exist_ok=True)
                filename = secure_filename(f"{novo_id}_{f.filename}")
                caminho = os.path.join(DATA_DIR, 'uploads', filename)
                f.save(caminho)
                novo['foto'] = filename

        materiais.append(novo)
        salvar_materiais(materiais)
        flash(f"{nome} adicionado ao estoque.")
        return redirect(url_for("estoque"))

    # dados usados pela busca (nome/gtin) que ajuda a evitar duplicados
    lista_busca = [
        {"id": m["id"], "nome": m["nome"], "gtin": m.get("gtin") or "", "quantidade": m["quantidade"], "unidade": m["unidade"]}
        for m in materiais
    ]
    # generate a one-time token to prevent duplicate form submits
    session['form_token'] = str(uuid.uuid4())
    return render_template(
        "adicionar.html",
        categorias=CATEGORIAS_EMOJI,
        unidades=UNIDADES,
        materiais_json=lista_busca,
        form_token=session['form_token'],
    )


@app.route("/baixa", methods=["GET", "POST"])
@requires_permission('baixa', 'create')
def baixa():
    materiais = carregar_materiais()

    if request.method == "POST":
        material_id = request.form.get("material_id", "").strip()
        if not material_id:
            flash("Selecione um material válido para dar baixa.")
            return redirect(url_for("baixa"))

        m = encontrar(materiais, material_id)
        if not m:
            flash("Material não encontrado no estoque. Ação cancelada.")
            return redirect(url_for("baixa"))

        qtd = parse_float_ptbr(request.form.get("quantidade", 0))
        motivo = request.form.get("motivo", "").strip()
        # Se selecionou "Outro" e digitou um motivo customizado, usar ele
        if motivo == "Outro":
            motivo_custom = request.form.get("motivo_custom", "").strip()
            if motivo_custom:
                motivo = motivo_custom

        unidade = (m.get("unidade") or "").lower()
        is_inteiro = unidade.startswith("unid") or unidade in ("unidades", "pares", "pacotes", "unidade", "unid")

        if is_inteiro:
            if abs(qtd - round(qtd)) > 1e-6:
                flash(f"Para o material '{m.get('nome')}' (unidade: {m.get('unidade')}), utilize apenas números inteiros.")
                return redirect(url_for("baixa"))
            qtd = int(round(qtd))

        if qtd <= 0:
            flash("Informe uma quantidade válida e maior que zero para a baixa.")
            return redirect(url_for("baixa"))

        # Checa estoque disponível
        current = float(m.get("quantidade") or 0)
        if current <= 0:
            flash(f"O material '{m.get('nome')}' está com estoque esgotado.")
            return redirect(url_for("baixa"))

        if qtd > current:
            flash(f"Quantidade insuficiente em estoque para {m.get('nome')}. Saldo atual: {current} {m.get('unidade')}.")
            return redirect(url_for("baixa"))

        # Executa dedução
        if is_inteiro:
            novo_q = max(0, int(current) - int(qtd))
        else:
            novo_q = round(max(0.0, current - float(qtd)), 3)

        m["quantidade"] = novo_q
        salvar_materiais(materiais)
        registrar_movimentacao("baixa", qtd, m.get("unidade"), motivo, m.get("nome"))
        flash(f"Baixa de {qtd} {m.get('unidade')} registrada em {m.get('nome')}.")
        return redirect(url_for("baixa"))

    mid_preselecionado = request.args.get("mid", "")
    return render_template(
        "baixa.html",
        materiais=materiais,
        motivos=MOTIVOS_BAIXA,
        mid_preselecionado=mid_preselecionado,
    )


# ── Produtos & Receitas ───────────────────────────────────────────────────────
DEFAULT_PRODUTOS = [
    {"nome": "Bolsa Tote Clássica", "emoji": "👜", "preco_venda": 180.0, "receita": [
        {"material_nome": "Courino Preto", "quantidade": 1.5},
        {"material_nome": "Zíper 30cm Preto", "quantidade": 1.0},
        {"material_nome": "Linha de Costura Preta", "quantidade": 0.05},
    ]},
    {"nome": "Necessaire Compacta", "emoji": "👝", "preco_venda": 60.0, "receita": [
        {"material_nome": "Courino Preto", "quantidade": 0.4},
        {"material_nome": "Zíper 30cm Preto", "quantidade": 1.0},
    ]},
    {"nome": "Bolsa Transversal Pequena", "emoji": "👛", "preco_venda": 120.0, "receita": [
        {"material_nome": "Courino Caramelo", "quantidade": 0.8},
        {"material_nome": "Mosquetão Dourado", "quantidade": 2.0},
        {"material_nome": "Fivela Quadrada Dourada", "quantidade": 1.0},
    ]},
]


def carregar_produtos():
    if USE_SQLITE:
        init_db()
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        cur.execute("SELECT COUNT(1) as cnt FROM produtos")
        cnt = cur.fetchone()[0]
        if cnt == 0:
            # Seed produtos padrão do sistema, resolvendo os materiais da receita por nome
            try:
                mat_map = {m["nome"].lower(): m for m in carregar_materiais()}
            except Exception:
                mat_map = {}
            now = agora().isoformat()
            for p in DEFAULT_PRODUTOS:
                receita = []
                for item in p.get("receita", []):
                    m = mat_map.get((item.get("material_nome") or "").lower())
                    if m:
                        receita.append({"material_id": m["id"], "quantidade": float(item.get("quantidade") or 0)})
                if not receita:
                    continue
                cur.execute(
                    "INSERT OR IGNORE INTO produtos (id, nome, emoji, preco_venda, receita, created_at, updated_at) VALUES (?, ?, ?, ?, ?, ?, ?)",
                    (str(uuid.uuid4()), p["nome"], p["emoji"], float(p["preco_venda"] or 0), json.dumps(receita, ensure_ascii=False), now, now),
                )
            conn.commit()
        cur.execute("SELECT * FROM produtos ORDER BY nome COLLATE NOCASE")
        rows = cur.fetchall()
        conn.close()
        res = []
        for r in rows:
            receita = []
            try:
                receita = json.loads(r["receita"]) if r["receita"] else []
            except Exception:
                receita = []
            res.append({
                "id": r["id"],
                "nome": r["nome"],
                "emoji": r["emoji"],
                "preco_venda": r["preco_venda"],
                "receita": receita,
                "gtin": r["gtin"] if "gtin" in r.keys() else "",
                "estoque_pronto": int(r["estoque_pronto"] or 0) if "estoque_pronto" in r.keys() else 0,
            })
        return res
    return carregar_json("produtos.json")


def salvar_produtos(produtos):
    if USE_SQLITE:
        init_db()
        conn = sqlite3.connect(DB_PATH)
        cur = conn.cursor()
        try:
            cur.execute("BEGIN IMMEDIATE")
            cur.execute("DELETE FROM produtos")
            now = agora().isoformat()
            for p in produtos:
                _id = p.get("id") or str(uuid.uuid4())
                cur.execute(
                    "INSERT INTO produtos (id,nome,emoji,preco_venda,receita,gtin,estoque_pronto,created_at,updated_at) VALUES (?,?,?,?,?,?,?,?,?)",
                    (_id, p.get("nome"), p.get("emoji"), float(p.get("preco_venda") or 0), json.dumps(p.get("receita") or [], ensure_ascii=False), p.get("gtin") or "", int(p.get("estoque_pronto") or 0), now, now),
                )
            conn.commit()
        except Exception:
            conn.rollback()
            raise
        finally:
            conn.close()
        return
    return salvar_json("produtos.json", produtos)


def carregar_pedidos():
    if USE_SQLITE:
        init_db()
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        cur.execute("SELECT * FROM pedidos ORDER BY data_pedido_iso DESC")
        rows = cur.fetchall()
        conn.close()
        res = []
        for r in rows:
            res.append({
                "id": r["id"],
                "cliente": r["cliente"],
                "produto_id": r["produto_id"],
                "produto_nome": r["produto_nome"],
                "produto_emoji": r["produto_emoji"],
                "quantidade": r["quantidade"],
                "preco_unitario": r["preco_unitario"],
                "valor_total": r["valor_total"],
                "status": r["status"],
                "materiais_baixados": bool(r["materiais_baixados"]),
                "usou_estoque_pronto": bool(r["usou_estoque_pronto"]) if "usou_estoque_pronto" in r.keys() else False,
                "data_pedido": r["data_pedido"],
                "data_pedido_iso": r["data_pedido_iso"],
                "observacoes": r["observacoes"],
            })
        return res
    lista = carregar_json("pedidos.json")
    return sorted(lista, key=lambda p: p.get("data_pedido_iso", ""), reverse=True)


def carregar_sobras():
    if USE_SQLITE:
        init_db()
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        cur.execute("SELECT * FROM sobras ORDER BY created_at DESC")
        rows = cur.fetchall()
        conn.close()
        sobs = []
        for r in rows:
            sobs.append({
                "id": r["id"],
                "material_id": r["material_id"],
                "descricao": r["descricao"],
                "quantidade": r["quantidade"],
                "unidade": r["unidade"],
                "data": r["data"],
                "status": r["status"],
            })
        return sobs
    lista = carregar_json("sobras.json")
    return list(reversed(lista))


def carregar_despesas():
    if USE_SQLITE:
        init_db()
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        cur.execute("SELECT * FROM despesas ORDER BY created_at DESC")
        rows = cur.fetchall()
        conn.close()
        return [{
            "id": r["id"],
            "descricao": r["descricao"],
            "valor": r["valor"],
            "categoria": r["categoria"],
            "data": r["data"],
            "created_at": r["created_at"] if "created_at" in r.keys() else ""
        } for r in rows]
    return carregar_json("despesas.json")


def carregar_movimentacoes(limit=100):
    if USE_SQLITE:
        init_db()
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        cur.execute("SELECT * FROM movimentacoes ORDER BY created_at DESC LIMIT ?", (limit,))
        rows = cur.fetchall()
        conn.close()
        return [{
            "id": r["id"],
            "tipo": r["tipo"],
            "material_nome": r["material_nome"],
            "quantidade": r["quantidade"],
            "unidade": r["unidade"],
            "motivo": r["motivo"],
            "data": r["data"],
            "usuario": r["usuario"],
        } for r in rows]
    return carregar_json("movimentacoes.json")[:limit]


# ── Relatórios Personalizados (CRUD & Engine de Gráficos) ──────────────────────
PALETA_CORES_GRAFICO = [
    "#7C3D12", "#C88242", "#D99B26", "#2E7D32", "#C62828",
    "#5C2D0E", "#A67C52", "#4A6B82", "#8D6E63", "#388E3C"
]


def carregar_relatorios_customizados():
    if USE_SQLITE:
        init_db()
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        cur.execute("SELECT * FROM relatorios_customizados ORDER BY created_at DESC")
        rows = cur.fetchall()
        conn.close()
        return [{
            "id": r["id"],
            "titulo": r["titulo"],
            "tipo": r["tipo"],
            "tipo_grafico": r["tipo_grafico"],
            "categoria_filtro": r["categoria_filtro"] or "",
            "status_filtro": r["status_filtro"] or "",
            "apenas_criticos": bool(r["apenas_criticos"]),
            "observacoes": r["observacoes"] or "",
            "criado_por": r["criado_por"] or "",
            "created_at": r["created_at"] or "",
            "updated_at": r["updated_at"] or "",
        } for r in rows]
    return carregar_json("relatorios_customizados.json")


def encontrar_relatorio_por_id(relatorio_id):
    lista = carregar_relatorios_customizados()
    for r in lista:
        if r["id"] == relatorio_id:
            return r
    return None


def salvar_relatorio_customizado(rel):
    _id = rel.get("id") or str(uuid.uuid4())
    now = agora().isoformat()
    if USE_SQLITE:
        init_db()
        conn = sqlite3.connect(DB_PATH)
        cur = conn.cursor()
        cur.execute("SELECT id FROM relatorios_customizados WHERE id=?", (_id,))
        exists = cur.fetchone()
        if exists:
            cur.execute("""
                UPDATE relatorios_customizados
                SET titulo=?, tipo=?, tipo_grafico=?, categoria_filtro=?, status_filtro=?,
                    apenas_criticos=?, observacoes=?, updated_at=?
                WHERE id=?
            """, (
                rel["titulo"], rel["tipo"], rel["tipo_grafico"], rel.get("categoria_filtro", ""),
                rel.get("status_filtro", ""), 1 if rel.get("apenas_criticos") else 0,
                rel.get("observacoes", ""), now, _id
            ))
        else:
            cur.execute("""
                INSERT INTO relatorios_customizados (id, titulo, tipo, tipo_grafico, categoria_filtro, status_filtro, apenas_criticos, observacoes, criado_por, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                _id, rel["titulo"], rel["tipo"], rel["tipo_grafico"], rel.get("categoria_filtro", ""),
                rel.get("status_filtro", ""), 1 if rel.get("apenas_criticos") else 0,
                rel.get("observacoes", ""), rel.get("criado_por", ""), now, now
            ))
        conn.commit()
        conn.close()
        rel["id"] = _id
        return rel

    lista = carregar_json("relatorios_customizados.json")
    idx = next((i for i, x in enumerate(lista) if x["id"] == _id), None)
    rel["id"] = _id
    rel["updated_at"] = now
    if idx is not None:
        lista[idx] = rel
    else:
        rel["created_at"] = now
        lista.append(rel)
    salvar_json("relatorios_customizados.json", lista)
    return rel


def excluir_relatorio_customizado(relatorio_id):
    if USE_SQLITE:
        init_db()
        conn = sqlite3.connect(DB_PATH)
        cur = conn.cursor()
        cur.execute("DELETE FROM relatorios_customizados WHERE id=?", (relatorio_id,))
        conn.commit()
        conn.close()
        return True
    lista = carregar_json("relatorios_customizados.json")
    lista = [r for r in lista if r["id"] != relatorio_id]
    salvar_json("relatorios_customizados.json", lista)
    return True


def gerar_dados_relatorio(relatorio):
    """Compila dinamicamente KPIs, dados tabulares e payloads Chart.js para qualquer relatório."""
    tipo = relatorio.get("tipo", "estoque")
    tipo_grafico = relatorio.get("tipo_grafico", "bar")
    cat_filtro = (relatorio.get("categoria_filtro") or "").strip()
    status_filtro = (relatorio.get("status_filtro") or "").strip()
    apenas_criticos = bool(relatorio.get("apenas_criticos"))

    kpis = []
    tabela = {"colunas": [], "linhas": []}
    chart_data = {"labels": [], "datasets": []}

    if tipo == "estoque":
        materiais = carregar_materiais()
        if cat_filtro:
            materiais = [m for m in materiais if m.get("categoria") == cat_filtro]
        if apenas_criticos:
            materiais = [m for m in materiais if m.get("quantidade", 0) <= m.get("quantidade_minima", 0)]

        total_itens = len(materiais)
        valor_total = round(sum(m.get("quantidade", 0) * m.get("custo", 0) for m in materiais), 2)
        criticos_count = len([m for m in materiais if m.get("quantidade", 0) <= m.get("quantidade_minima", 0)])

        kpis = [
            {"titulo": "Itens Filtrados", "valor": f"{total_itens}", "sub": "Materiais selecionados", "cor": "var(--primary)"},
            {"titulo": "Valor em Estoque", "valor": f"R$ {formatar_reais(valor_total)}", "sub": "Total imobilizado", "cor": "var(--accent)"},
            {"titulo": "Nível Crítico", "valor": f"{criticos_count}", "sub": "Itens abaixo do mínimo", "cor": "var(--danger)" if criticos_count > 0 else "var(--success)"},
        ]

        tabela["colunas"] = ["Material", "Categoria", "Qtd Atual", "Qtd Mínima", "Custo Unit.", "Valor Total", "Status"]
        tabela["linhas"] = []
        labels = []
        data_qtd = []
        data_val = []

        for m in materiais:
            val_t = round(m.get("quantidade", 0) * m.get("custo", 0), 2)
            is_crit = m.get("quantidade", 0) <= m.get("quantidade_minima", 0)
            tabela["linhas"].append([
                f"{m.get('emoji','')} {m.get('nome','')}",
                m.get("categoria", ""),
                f"{m.get('quantidade',0)} {m.get('unidade','')}",
                f"{m.get('quantidade_minima',0)} {m.get('unidade','')}",
                f"R$ {formatar_reais(float(m.get('custo',0)))}",
                f"R$ {formatar_reais(val_t)}",
                "⚠️ Crítico" if is_crit else "✅ Normal"
            ])
            labels.append(m.get("nome", "")[:18])
            data_qtd.append(m.get("quantidade", 0))
            data_val.append(val_t)

        chart_data["labels"] = labels
        if tipo_grafico in ("doughnut", "pie", "polarArea"):
            chart_data["datasets"] = [{
                "label": "Valor em Estoque (R$)",
                "data": data_val,
                "backgroundColor": PALETA_CORES_GRAFICO[:len(labels)],
                "borderWidth": 1.5,
                "borderColor": "#FFFFFF"
            }]
        else:
            chart_data["datasets"] = [
                {
                    "label": "Qtd Atual",
                    "data": data_qtd,
                    "backgroundColor": "rgba(124, 61, 18, 0.8)",
                    "borderColor": "#7C3D12",
                    "borderWidth": 1.5,
                },
                {
                    "label": "Valor Total (R$)",
                    "data": data_val,
                    "backgroundColor": "rgba(200, 130, 66, 0.8)",
                    "borderColor": "#C88242",
                    "borderWidth": 1.5,
                }
            ]

    elif tipo == "financeiro":
        despesas = carregar_despesas()
        pedidos = carregar_pedidos()
        materiais = carregar_materiais()

        if cat_filtro:
            despesas = [d for d in despesas if d.get("categoria") == cat_filtro]

        total_desp = round(sum(d.get("valor", 0) for d in despesas), 2)
        rec_entregue = round(sum(p.get("valor_total", 0) for p in pedidos if p.get("status") == "Entregue"), 2)
        lucro = round(rec_entregue - total_desp, 2)

        kpis = [
            {"titulo": "Receita Recebida", "valor": f"R$ {formatar_reais(rec_entregue)}", "sub": "Pedidos entregues", "cor": "var(--success)"},
            {"titulo": "Despesas Filtradas", "valor": f"R$ {formatar_reais(total_desp)}", "sub": f"{len(despesas)} lançamentos", "cor": "var(--danger)"},
            {"titulo": "Lucro Realizado", "valor": f"R$ {formatar_reais(lucro)}", "sub": "Receita − Despesas", "cor": "var(--success)" if lucro >= 0 else "var(--danger)"},
        ]

        tabela["colunas"] = ["Data", "Descrição", "Categoria", "Valor (R$)"]
        tabela["linhas"] = [[d.get("data",""), d.get("descricao",""), d.get("categoria","Outros"), f"R$ {formatar_reais(float(d.get('valor',0)))}"] for d in despesas]

        # Agrupamento de despesas por categoria
        desp_por_cat = {}
        for d in despesas:
            c = d.get("categoria", "Outros")
            desp_por_cat[c] = round(desp_por_cat.get(c, 0.0) + float(d.get("valor", 0)), 2)

        if tipo_grafico in ("doughnut", "pie", "polarArea"):
            chart_data["labels"] = list(desp_por_cat.keys())
            chart_data["datasets"] = [{
                "label": "Despesas por Categoria (R$)",
                "data": list(desp_por_cat.values()),
                "backgroundColor": PALETA_CORES_GRAFICO[:len(desp_por_cat)],
                "borderWidth": 1.5,
                "borderColor": "#FFFFFF"
            }]
        else:
            chart_data["labels"] = ["Receita Recebida", "Despesas Totais", "Lucro Realizado"]
            chart_data["datasets"] = [{
                "label": "Comparativo Financeiro (R$)",
                "data": [rec_entregue, total_desp, max(0, lucro)],
                "backgroundColor": ["rgba(46, 125, 50, 0.8)", "rgba(198, 40, 40, 0.8)", "rgba(200, 130, 66, 0.8)"],
                "borderColor": ["#2E7D32", "#C62828", "#C88242"],
                "borderWidth": 1.5,
            }]

    elif tipo == "pedidos":
        pedidos = carregar_pedidos()
        if status_filtro:
            pedidos = [p for p in pedidos if p.get("status") == status_filtro]

        total_peds = len(pedidos)
        faturamento_tot = round(sum(p.get("valor_total", 0) for p in pedidos), 2)
        entregues = len([p for p in pedidos if p.get("status") == "Entregue"])

        kpis = [
            {"titulo": "Total de Pedidos", "valor": f"{total_peds}", "sub": "Registros filtrados", "cor": "var(--primary)"},
            {"titulo": "Faturamento", "valor": f"R$ {formatar_reais(faturamento_tot)}", "sub": "Volume total", "cor": "var(--accent)"},
            {"titulo": "Entregues", "valor": f"{entregues}", "sub": f"{round((entregues/total_peds*100) if total_peds>0 else 0)}% do total", "cor": "var(--success)"},
        ]

        tabela["colunas"] = ["Cliente", "Produto", "Qtd", "Valor Total", "Status", "Data Pedido"]
        tabela["linhas"] = [
            [p.get("cliente",""), p.get("produto_nome",""), p.get("quantidade",1), f"R$ {formatar_reais(float(p.get('valor_total',0)))}", p.get("status","Pendente"), p.get("data_pedido","")]
            for p in pedidos
        ]

        # Agrupamento por status
        ped_por_status = {}
        for p in pedidos:
            st = p.get("status", "Pendente")
            ped_por_status[st] = ped_por_status.get(st, 0) + 1

        chart_data["labels"] = list(ped_por_status.keys())
        chart_data["datasets"] = [{
            "label": "Qtd de Pedidos",
            "data": list(ped_por_status.values()),
            "backgroundColor": PALETA_CORES_GRAFICO[:len(ped_por_status)],
            "borderWidth": 1.5,
            "borderColor": "#FFFFFF"
        }]

    elif tipo == "movimentacoes":
        movs = carregar_movimentacoes(150)
        tabela["colunas"] = ["Data", "Tipo", "Material", "Qtd", "Motivo", "Usuário"]
        tabela["linhas"] = [
            [m.get("data",""), m.get("tipo",""), m.get("material_nome",""), f"{m.get('quantidade',0)} {m.get('unidade','')}", m.get("motivo",""), m.get("usuario","")]
            for m in movs
        ]

        # Agrupamento por tipo
        por_tipo = {}
        for m in movs:
            t = m.get("tipo", "Outro")
            por_tipo[t] = por_tipo.get(t, 0) + 1

        kpis = [
            {"titulo": "Total Movimentações", "valor": f"{len(movs)}", "sub": "Últimos registros", "cor": "var(--primary)"},
            {"titulo": "Saídas / Baixas", "valor": f"{por_tipo.get('Saída', por_tipo.get('Baixa', 0))}", "sub": "Consumo de produção", "cor": "var(--danger)"},
            {"titulo": "Entradas", "valor": f"{por_tipo.get('Entrada', 0)}", "sub": "Reposição de estoque", "cor": "var(--success)"},
        ]

        chart_data["labels"] = list(por_tipo.keys())
        chart_data["datasets"] = [{
            "label": "Movimentações",
            "data": list(por_tipo.values()),
            "backgroundColor": PALETA_CORES_GRAFICO[:len(por_tipo)],
            "borderWidth": 1.5,
            "borderColor": "#FFFFFF"
        }]

    elif tipo == "sobras":
        sobras = carregar_sobras()
        if status_filtro:
            sobras = [s for s in sobras if s.get("status") == status_filtro]

        total_sobras = len(sobras)
        disp = len([s for s in sobras if s.get("status") == "Disponível"])
        util = len([s for s in sobras if s.get("status") == "Utilizado"])

        kpis = [
            {"titulo": "Total de Sobras", "valor": f"{total_sobras}", "sub": "Cadastradas", "cor": "var(--primary)"},
            {"titulo": "Disponíveis", "valor": f"{disp}", "sub": "Prontas para uso", "cor": "var(--accent)"},
            {"titulo": "Reaproveitadas", "valor": f"{util}", "sub": "Peças geradas", "cor": "var(--success)"},
        ]

        tabela["colunas"] = ["Descrição", "Quantidade", "Data", "Status"]
        tabela["linhas"] = [
            [s.get("descricao",""), f"{s.get('quantidade',0)} {s.get('unidade','')}", s.get("data",""), s.get("status","Disponível")]
            for s in sobras
        ]

        chart_data["labels"] = ["Disponível", "Utilizado", "Descartado"]
        chart_data["datasets"] = [{
            "label": "Status das Sobras",
            "data": [disp, util, total_sobras - (disp + util)],
            "backgroundColor": ["#C88242", "#2E7D32", "#7A6B63"],
            "borderWidth": 1.5,
            "borderColor": "#FFFFFF"
        }]

    else: # Geral
        materiais = carregar_materiais()
        pedidos = carregar_pedidos()
        despesas = carregar_despesas()

        val_est = round(sum(m.get("quantidade", 0) * m.get("custo", 0) for m in materiais), 2)
        rec_ent = round(sum(p.get("valor_total", 0) for p in pedidos if p.get("status") == "Entregue"), 2)
        rec_prev = round(sum(p.get("valor_total", 0) for p in pedidos if p.get("status") != "Entregue"), 2)
        tot_desp = round(sum(d.get("valor", 0) for d in despesas), 2)
        lucro = round(rec_ent - tot_desp, 2)

        kpis = [
            {"titulo": "Valor em Estoque", "valor": f"R$ {formatar_reais(val_est)}", "sub": f"{len(materiais)} insumos", "cor": "var(--primary)"},
            {"titulo": "Receita Recebida", "valor": f"R$ {formatar_reais(rec_ent)}", "sub": "Pedidos entregues", "cor": "var(--success)"},
            {"titulo": "Lucro Operacional", "valor": f"R$ {formatar_reais(lucro)}", "sub": "Receita − Despesas", "cor": "var(--success)" if lucro >= 0 else "var(--danger)"},
        ]

        tabela["colunas"] = ["Métrica Consolidada", "Valor"]
        tabela["linhas"] = [
            ["Valor Total em Estoque", f"R$ {formatar_reais(val_est)}"],
            ["Receita Recebida", f"R$ {formatar_reais(rec_ent)}"],
            ["Receita Prevista (Em andamento)", f"R$ {formatar_reais(rec_prev)}"],
            ["Despesas Operacionais", f"R$ {formatar_reais(tot_desp)}"],
            ["Lucro Líquido", f"R$ {formatar_reais(lucro)}"],
        ]

        chart_data["labels"] = ["Estoque", "Receita Entregue", "Receita Prevista", "Despesas", "Lucro"]
        chart_data["datasets"] = [{
            "label": "Balanço Geral (R$)",
            "data": [val_est, rec_ent, rec_prev, tot_desp, max(0, lucro)],
            "backgroundColor": ["#7C3D12", "#2E7D32", "#C88242", "#C62828", "#5C2D0E"],
            "borderWidth": 1.5,
            "borderColor": "#FFFFFF"
        }]

    return {
        "kpis": kpis,
        "tabela": tabela,
        "chart_data": chart_data,
        "tipo_grafico": tipo_grafico if tipo_grafico in ("bar", "line", "doughnut", "pie", "polarArea") else "bar"
    }


def calcular_produto(produto, mat_map):
    """Anexa custo estimado, margem e a receita já resolvida com nome/emoji/unidade dos materiais."""
    custo = 0.0
    receita_detalhada = []
    for item in produto.get("receita", []):
        m = mat_map.get(item["material_id"])
        if m:
            custo += m["custo"] * item["quantidade"]
            receita_detalhada.append({
                "nome": m["nome"], "emoji": m["emoji"],
                "quantidade": item["quantidade"], "unidade": m["unidade"],
                "disponivel": m["quantidade"],
            })
        else:
            receita_detalhada.append({
                "nome": "Material removido", "emoji": "❓",
                "quantidade": item["quantidade"], "unidade": "", "disponivel": None,
            })
    p = dict(produto)
    p["custo_estimado"] = round(custo, 2)
    p["margem"] = round(produto.get("preco_venda", 0) - custo, 2)
    p["receita_detalhada"] = receita_detalhada
    return p


@app.route("/produtos")
@requires_permission('produtos', 'read')
def produtos():
    lista = carregar_produtos()
    mat_map = {m["id"]: m for m in carregar_materiais()}
    produtos_calc = [calcular_produto(p, mat_map) for p in lista]
    return render_template("produtos.html", produtos=produtos_calc)


@app.route("/produtos/novo", methods=["GET", "POST"])
@requires_permission('produtos', 'create')
def produto_novo():
    materiais = carregar_materiais()

    if request.method == "POST":
        nome = request.form.get("nome", "").strip()
        emoji = request.form.get("emoji", "👜").strip() or "👜"
        preco_venda = parse_float_ptbr(request.form.get("preco_venda", 0))

        mat_ids = request.form.getlist("material_id[]")
        qtds = request.form.getlist("material_qtd[]")
        receita = []
        for mid, q in zip(mat_ids, qtds):
            if not mid:
                continue
            qf = parse_float_ptbr(q)
            if qf <= 0:
                continue
            receita.append({"material_id": mid, "quantidade": qf})

        if not nome:
            flash("Informe o nome do produto.")
            return redirect(url_for("produto_novo"))

        # must have at least one material in receita
        if not receita:
            flash("Não é possível criar um produto sem materiais na receita.")
            return redirect(url_for("produto_novo"))

        # prevent duplicate product names (case-insensitive)
        produtos_lista = carregar_produtos()
        if any(p.get('nome','').strip().lower() == nome.lower() for p in produtos_lista):
            flash('Já existe um produto com este nome.')
            return redirect(url_for('produto_novo'))

        gtin = request.form.get("gtin", "").strip()
        estoque_pronto = int(parse_float_ptbr(request.form.get("estoque_pronto", 0)))

        produtos_lista.append({
            "id": str(uuid.uuid4()),
            "nome": nome,
            "emoji": emoji,
            "preco_venda": preco_venda,
            "receita": receita,
            "gtin": gtin,
            "estoque_pronto": max(0, estoque_pronto),
        })
        salvar_produtos(produtos_lista)
        flash(f"{nome} cadastrado em Produtos & Receitas.")
        return redirect(url_for("produtos"))

    return render_template("produto_form.html", materiais=materiais, emojis=EMOJIS_PRODUTO)


@app.route("/produtos/<produto_id>/ajuste_estoque", methods=["POST"])
@requires_permission('produtos', 'update')
def produto_ajuste_estoque(produto_id):
    produtos_lista = carregar_produtos()
    p = next((prod for prod in produtos_lista if prod["id"] == produto_id), None)
    if not p:
        flash("Produto não encontrado.")
        return redirect(url_for("produtos"))

    acao = request.form.get("acao", "adicionar").strip()
    qtd = int(parse_float_ptbr(request.form.get("quantidade", 1)))
    if qtd <= 0:
        qtd = 1

    atual = int(p.get("estoque_pronto") or 0)
    now = agora()
    now_str = now.strftime("%d/%m/%Y %H:%M")
    usuario_id = session.get("user_id") if session else None

    if acao == "adicionar":
        novo = atual + qtd
        motivo = request.form.get("motivo", "").strip() or "Entrada manual no estoque de peças prontas"
        p["estoque_pronto"] = novo
        salvar_produtos(produtos_lista)
        registrar_movimentacao("estoque_pronto", qtd, "unidades", motivo, p["nome"])
        flash(f"+{qtd} peça(s) de {p['nome']} adicionada(s) ao estoque de pronta-entrega (Total: {novo}).")
    elif acao == "remover":
        novo = max(0, atual - qtd)
        p["estoque_pronto"] = novo
        salvar_produtos(produtos_lista)
        registrar_movimentacao("estoque_pronto", qtd, "unidades", "Saída manual do estoque de peças prontas", p["nome"])
        flash(f"-{qtd} peça(s) de {p['nome']} retirada(s) do estoque de pronta-entrega (Total: {novo}).")

    return redirect(url_for("produtos"))


@app.route("/produtos/<produto_id>/excluir", methods=["POST"])
@requires_permission('produtos', 'delete')
def produto_excluir(produto_id):
    produtos_lista = carregar_produtos()
    produtos_lista = [p for p in produtos_lista if p["id"] != produto_id]
    salvar_produtos(produtos_lista)
    flash("Produto removido.")
    return redirect(url_for("produtos"))


# ── Pedidos dos Clientes ──────────────────────────────────────────────────────
@app.route("/pedidos")
@requires_permission('pedidos', 'read')
def pedidos():
    if USE_SQLITE:
        init_db()
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        cur.execute("SELECT * FROM pedidos ORDER BY data_pedido_iso DESC")
        rows = cur.fetchall()
        conn.close()
        res = []
        for r in rows:
            res.append({
                "id": r["id"],
                "cliente": r["cliente"],
                "produto_id": r["produto_id"],
                "produto_nome": r["produto_nome"],
                "produto_emoji": r["produto_emoji"],
                "quantidade": r["quantidade"],
                "preco_unitario": r["preco_unitario"],
                "valor_total": r["valor_total"],
                "status": r["status"],
                "materiais_baixados": bool(r["materiais_baixados"]),
                "data_pedido": r["data_pedido"],
                "data_pedido_iso": r["data_pedido_iso"],
                "observacoes": r["observacoes"],
            })
        return render_template("pedidos.html", pedidos=res, status_lista=STATUS_PEDIDO, status_badge=STATUS_PEDIDO_BADGE)

    lista = carregar_json("pedidos.json")
    lista_ordenada = sorted(lista, key=lambda p: p.get("data_pedido_iso", ""), reverse=True)
    return render_template("pedidos.html", pedidos=lista_ordenada, status_lista=STATUS_PEDIDO,
                            status_badge=STATUS_PEDIDO_BADGE)


@app.route("/pedidos/novo", methods=["GET", "POST"])
@requires_permission('pedidos', 'create')
def pedido_novo():
    produtos_lista = carregar_produtos() if USE_SQLITE else carregar_json("produtos.json")

    if request.method == "POST":
        cliente = request.form.get("cliente", "").strip()
        produto_id = request.form.get("produto_id", "")
        try:
            quantidade = int(request.form.get("quantidade", 1) or 1)
        except ValueError:
            quantidade = 1
        observacoes = request.form.get("observacoes", "").strip()

        produto = next((p for p in produtos_lista if p["id"] == produto_id), None)
        if not cliente or not produto or quantidade <= 0:
            flash("Preencha cliente, produto e uma quantidade válida.")
            return redirect(url_for("pedido_novo"))

        estoque_pronto_atual = int(produto.get("estoque_pronto") or 0)
        usar_pronta = estoque_pronto_atual >= quantidade

        preco_unit = produto.get("preco_venda", 0)
        dt_pedido = agora()
        novo = {
            "id": str(uuid.uuid4()),
            "cliente": cliente,
            "produto_id": produto_id,
            "produto_nome": produto["nome"],
            "produto_emoji": produto.get("emoji", "👜"),
            "quantidade": quantidade,
            "preco_unitario": preco_unit,
            "valor_total": round(preco_unit * quantidade, 2),
            "status": "Concluído" if usar_pronta else "Pendente",
            "materiais_baixados": 1 if usar_pronta else 0,
            "usou_estoque_pronto": 1 if usar_pronta else 0,
            "data_pedido": dt_pedido.strftime("%d/%m/%Y"),
            "data_pedido_iso": dt_pedido.strftime("%Y-%m-%d %H:%M:%S"),
            "observacoes": observacoes,
        }
        if USE_SQLITE:
            init_db()
            conn = sqlite3.connect(DB_PATH)
            cur = conn.cursor()
            if usar_pronta:
                cur.execute("UPDATE produtos SET estoque_pronto=?, updated_at=? WHERE id=?", (estoque_pronto_atual - quantidade, dt_pedido.isoformat(), produto_id))
                cur.execute(
                    "INSERT INTO movimentacoes (id, tipo, material_nome, quantidade, unidade, motivo, data, usuario, created_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
                    (
                        str(uuid.uuid4()),
                        "estoque_pronto",
                        produto["nome"],
                        quantidade,
                        "unidades",
                        f"Atendimento de pedido de {cliente} usando peça pronta em estoque",
                        dt_pedido.strftime("%d/%m/%Y %H:%M"),
                        session.get("user_id"),
                        dt_pedido.isoformat()
                    )
                )
            cur.execute(
                "INSERT INTO pedidos (id,cliente,produto_id,produto_nome,produto_emoji,quantidade,preco_unitario,valor_total,status,materiais_baixados,usou_estoque_pronto,data_pedido,data_pedido_iso,observacoes,created_at,updated_at) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                (
                    novo["id"], novo["cliente"], novo["produto_id"], novo["produto_nome"], novo["produto_emoji"], novo["quantidade"], novo["preco_unitario"], novo["valor_total"], novo["status"], novo["materiais_baixados"], novo["usou_estoque_pronto"], novo["data_pedido"], novo["data_pedido_iso"], novo["observacoes"], dt_pedido.isoformat(), dt_pedido.isoformat()
                )
            )
            conn.commit()
            conn.close()
        else:
            if usar_pronta:
                produto["estoque_pronto"] = estoque_pronto_atual - quantidade
                salvar_produtos(produtos_lista)
            pedidos_lista = carregar_json("pedidos.json")
            pedidos_lista.append(novo)
            salvar_json("pedidos.json", pedidos_lista)

        if usar_pronta:
            flash(f"Pedido de {cliente} registrado e atendido imediatamente com {quantidade} peça(s) pronta(s) do estoque!")
        else:
            flash(f"Pedido de {cliente} registrado.")
        return redirect(url_for("pedidos"))

    return render_template("pedido_form.html", produtos=produtos_lista)


@app.route("/pedidos/<pedido_id>/status", methods=["POST"])
@requires_permission('pedidos', 'update')
def pedido_status(pedido_id):
    novo_status = request.form.get("status", "").strip()
    if novo_status not in STATUS_PEDIDO:
        flash("Status inválido.")
        return redirect(url_for("pedidos"))

    now = agora()
    now_iso = now.isoformat()
    now_str = now.strftime("%d/%m/%Y %H:%M")
    usuario_id = session.get("user_id") if session else None

    if USE_SQLITE:
        init_db()
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        try:
            cur.execute("BEGIN IMMEDIATE")
            cur.execute("SELECT * FROM pedidos WHERE id=?", (pedido_id,))
            p = cur.fetchone()
            if not p:
                conn.rollback()
                flash("Pedido não encontrado.")
                return redirect(url_for("pedidos"))

            # Se o pedido for cancelado e já houve materiais baixados ou peça pronta usada, armazena a bolsa no estoque de pronta-entrega
            if novo_status == "Cancelado" and bool(p["materiais_baixados"]):
                cur.execute("SELECT estoque_pronto FROM produtos WHERE id=?", (p["produto_id"],))
                row_prod = cur.fetchone()
                est_atual = int(row_prod[0] or 0) if row_prod else 0
                qtd_ped = int(p["quantidade"] or 1)
                cur.execute("UPDATE produtos SET estoque_pronto=?, updated_at=? WHERE id=?", (est_atual + qtd_ped, now_iso, p["produto_id"]))
                cur.execute(
                    "INSERT INTO movimentacoes (id, tipo, material_nome, quantidade, unidade, motivo, data, usuario, created_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
                    (
                        str(uuid.uuid4()),
                        "estoque_pronto",
                        p["produto_nome"],
                        qtd_ped,
                        "unidades",
                        f"Bolsa armazenada no estoque de pronta-entrega após cancelamento do pedido de {p['cliente']}",
                        now_str,
                        usuario_id,
                        now_iso
                    )
                )
                cur.execute("UPDATE pedidos SET status=?, updated_at=? WHERE id=?", (novo_status, now_iso, pedido_id))
                conn.commit()
                flash(f'Pedido de {p["cliente"]} cancelado. {qtd_ped}x {p["produto_nome"]} foi guardada no estoque de peças prontas para outro cliente.')
                return redirect(url_for("pedidos"))

            # Se for mover para produção, concluído ou entregue e ainda não deu baixa automática dos materiais da receita
            if novo_status in ("Em produção", "Concluído", "Entregue") and not bool(p["materiais_baixados"]):
                cur.execute("SELECT * FROM produtos WHERE id=?", (p["produto_id"],))
                pr = cur.fetchone()
                if pr and pr["receita"]:
                    receita = []
                    try:
                        receita = json.loads(pr["receita"]) if isinstance(pr["receita"], str) else (pr["receita"] or [])
                    except Exception:
                        receita = []

                    if not isinstance(receita, list):
                        receita = []

                    # Pre-check disponibilidade dos materiais
                    insufficient = []
                    missing = []
                    qtd_pedido = float(p["quantidade"] or 1)

                    for item in receita:
                        mat_id = item.get("material_id")
                        qtd_por_unidade = float(item.get("quantidade") or 0)
                        total = round(qtd_por_unidade * qtd_pedido, 3)
                        cur.execute("SELECT quantidade, unidade, nome FROM materiais WHERE id=?", (mat_id,))
                        mat = cur.fetchone()
                        if not mat:
                            missing.append(str(mat_id))
                        else:
                            unidade = (mat["unidade"] or "").lower()
                            if unidade in ("unidades", "unidade", "unid") and abs(total - int(total)) > 1e-9:
                                insufficient.append(f"{mat['nome']}: quantidade precisa ser inteira (calculada {total})")
                            elif float(mat["quantidade"] or 0) < total:
                                insufficient.append(f"{mat['nome']}: estoque insuficiente ({mat['quantidade']} < {total})")

                    if missing or insufficient:
                        msg_parts = []
                        if missing:
                            msg_parts.append('Materiais ausentes: ' + ', '.join(missing))
                        if insufficient:
                            msg_parts.extend(insufficient)
                        conn.rollback()
                        flash('Não foi possível concluir a baixa automática: ' + '; '.join(msg_parts))
                        return redirect(url_for('pedidos'))

                    # Executa as deduções de materiais e registra movimentações na mesma transação atômica
                    for item in receita:
                        mat_id = item.get("material_id")
                        qtd_por_unidade = float(item.get("quantidade") or 0)
                        total = round(qtd_por_unidade * qtd_pedido, 3)
                        cur.execute("SELECT quantidade, unidade, nome FROM materiais WHERE id=?", (mat_id,))
                        mat = cur.fetchone()
                        if mat:
                            nova_qtd = round(max(0.0, float(mat["quantidade"] or 0) - total), 3)
                            cur.execute("UPDATE materiais SET quantidade=?, updated_at=? WHERE id=?", (nova_qtd, now_iso, mat_id))
                            cur.execute(
                                "INSERT INTO movimentacoes (id, tipo, material_nome, quantidade, unidade, motivo, data, usuario, created_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
                                (
                                    str(uuid.uuid4()),
                                    "producao",
                                    mat["nome"],
                                    total,
                                    mat["unidade"],
                                    f"Produção — pedido de {p['cliente']}",
                                    now_str,
                                    usuario_id,
                                    now_iso
                                )
                            )

                    cur.execute("UPDATE pedidos SET materiais_baixados=1, status=?, updated_at=? WHERE id=?", (novo_status, now_iso, pedido_id))
            else:
                cur.execute("UPDATE pedidos SET status=?, updated_at=? WHERE id=?", (novo_status, now_iso, pedido_id))

            conn.commit()
            flash(f'Pedido de {p["cliente"]} atualizado para "{novo_status}".')
        except Exception as e:
            try:
                conn.rollback()
            except Exception:
                pass
            flash("Erro ao atualizar status do pedido: " + str(e))
        finally:
            try:
                conn.close()
            except Exception:
                pass
        return redirect(url_for("pedidos"))

    # Legacy JSON path
    pedidos_lista = carregar_json("pedidos.json")
    pedido = next((p for p in pedidos_lista if p["id"] == pedido_id), None)
    if not pedido:
        flash("Pedido não encontrado.")
        return redirect(url_for("pedidos"))

    pedido["status"] = novo_status
    if novo_status in ("Em produção", "Concluído", "Entregue") and not pedido.get("materiais_baixados"):
        produtos_lista = carregar_json("produtos.json")
        produto = next((pr for pr in produtos_lista if pr["id"] == pedido["produto_id"]), None)
        if produto and produto.get("receita"):
            materiais = carregar_materiais()
            missing = []
            insufficient = []
            qtd_pedido = float(pedido.get("quantidade") or 1)
            for item in produto["receita"]:
                m = encontrar(materiais, item["material_id"])
                total = round(item["quantidade"] * qtd_pedido, 3)
                if not m:
                    missing.append(item.get("material_id"))
                else:
                    unidade = (m.get("unidade") or "").lower()
                    if unidade in ("unidades", "unidade", "unid") and abs(total - int(total)) > 1e-9:
                        insufficient.append(f"{m.get('nome')}: quantidade precisa ser inteira (calculada {total})")
                    elif float(m.get("quantidade") or 0) < total:
                        insufficient.append(f"{m.get('nome')}: estoque insuficiente ({m.get('quantidade')} < {total})")
            if missing or insufficient:
                msg = []
                if missing:
                    msg.append('Materiais ausentes: ' + ', '.join(missing))
                if insufficient:
                    msg.extend(insufficient)
                flash('Não foi possível concluir a baixa automática: ' + '; '.join(msg))
                return redirect(url_for("pedidos"))
            else:
                for item in produto["receita"]:
                    m = encontrar(materiais, item["material_id"])
                    total = round(item["quantidade"] * qtd_pedido, 3)
                    if m:
                        m["quantidade"] = round(max(0, m["quantidade"] - total), 3)
                        registrar_movimentacao("producao", total, m["unidade"], f"Produção — pedido de {pedido['cliente']}", m["nome"])
                salvar_materiais(materiais)
                pedido["materiais_baixados"] = True

    salvar_json("pedidos.json", pedidos_lista)
    flash(f'Pedido de {pedido["cliente"]} atualizado para "{novo_status}".')
    return redirect(url_for("pedidos"))


@app.route("/pedidos/<pedido_id>/excluir", methods=["POST"])
@requires_permission('pedidos', 'delete')
def pedido_excluir(pedido_id):
    if USE_SQLITE:
        init_db()
        conn = sqlite3.connect(DB_PATH)
        cur = conn.cursor()
        cur.execute("DELETE FROM pedidos WHERE id=?", (pedido_id,))
        conn.commit()
        conn.close()
        flash("Pedido removido.")
        return redirect(url_for("pedidos"))

    pedidos_lista = carregar_json("pedidos.json")
    pedidos_lista = [p for p in pedidos_lista if p["id"] != pedido_id]
    salvar_json("pedidos.json", pedidos_lista)
    flash("Pedido removido.")
    return redirect(url_for("pedidos"))


# ── Sobras e Reaproveitamento ──────────────────────────────────────────────────
@app.route("/sobras")
@requires_permission('sobras', 'read')
def sobras():
    return render_template("sobras.html", sobras=carregar_sobras(), status_badge=STATUS_SOBRA_BADGE)


@app.route("/sobras/novo", methods=["GET", "POST"])
@requires_permission('sobras', 'create')
def sobra_novo():
    materiais = carregar_materiais()

    if request.method == "POST":
        material_id = request.form.get("material_id", "").strip()
        descricao = request.form.get("descricao", "").strip()
        tipo_entrada = request.form.get("tipo_entrada", "direto").strip()

        m = encontrar(materiais, material_id) if material_id else None

        if tipo_entrada == "dimensoes":
            comp = parse_float_ptbr(request.form.get("comprimento_cm", 0))
            larg = parse_float_ptbr(request.form.get("largura_cm", 0))
            if comp <= 0 or larg <= 0:
                flash("Informe comprimento e largura válidos em centímetros (maiores que zero).")
                return redirect(url_for("sobra_novo"))

            area_m2 = round((comp * larg) / 10000.0, 4)
            # Para tecidos/courino em rolo padrão (1,40m largura), calcula fração correspondente
            equiv_metros = max(0.001, round(area_m2 / 1.40, 3))
            quantidade = equiv_metros
            unidade = "metros"

            if not descricao:
                m_nome = m["nome"] if m else "Material"
                descricao = f"Retalho de {m_nome} {comp:.0f}x{larg:.0f} cm ({area_m2:.2f} m²)"
        else:
            quantidade = parse_float_ptbr(request.form.get("quantidade", 0))
            unidade = request.form.get("unidade", "unidades")

        if not descricao:
            descricao = (m["nome"] + " (Sobra)") if m else "Sobra sem descrição"

        if quantidade <= 0:
            flash("Informe uma quantidade válida para a sobra.")
            return redirect(url_for("sobra_novo"))

        # Regra: se informou material_id, ele OBRIGATORIAMENTE precisa existir no estoque
        if material_id:
            if not m:
                flash("Não foi possível registrar a sobra: o material selecionado não existe no estoque.")
                return redirect(url_for("sobra_novo"))

            unidade_mat = (m.get("unidade") or "unidades").lower()
            is_inteiro = unidade_mat.startswith("unid") or unidade_mat in ("unidades", "pares", "pacotes")
            if is_inteiro:
                if abs(quantidade - round(quantidade)) > 1e-6:
                    flash(f"Para o material '{m.get('nome')}' (unidade: {m.get('unidade')}), utilize apenas números inteiros.")
                    return redirect(url_for("sobra_novo"))
                quantidade = int(round(quantidade))
                unidade = m.get("unidade", "unidades")

            saldo = float(m.get("quantidade") or 0)
            if quantidade > saldo:
                flash(f"Não foi possível registrar a sobra: estoque insuficiente para '{m.get('nome')}'. Saldo atual: {saldo} {m.get('unidade')}.")
                return redirect(url_for("sobra_novo"))

        now = agora()
        now_iso = now.isoformat()
        now_str = now.strftime("%d/%m/%Y %H:%M")

        if USE_SQLITE:
            init_db()
            conn = sqlite3.connect(DB_PATH)
            cur = conn.cursor()
            try:
                cur.execute("BEGIN IMMEDIATE")
                sobra_id = str(uuid.uuid4())
                cur.execute(
                    "INSERT INTO sobras (id,material_id,descricao,quantidade,unidade,data,status,created_at,updated_at) VALUES (?,?,?,?,?,?,?,?,?)",
                    (sobra_id, material_id or None, descricao, quantidade, unidade, now.strftime("%d/%m/%Y"), "Disponível", now_iso, now_iso)
                )
                if material_id and m:
                    saldo_mat = float(m.get("quantidade") or 0)
                    unidade_mat = (m.get("unidade") or "").lower()
                    is_inteiro = unidade_mat.startswith("unid") or unidade_mat in ("unidades", "pares", "pacotes")
                    nova_qtd = max(0, int(saldo_mat) - int(quantidade)) if is_inteiro else round(max(0.0, saldo_mat - float(quantidade)), 3)
                    cur.execute(
                        "UPDATE materiais SET quantidade=?, updated_at=? WHERE id=?",
                        (nova_qtd, now_iso, material_id)
                    )
                    cur.execute(
                        "INSERT INTO movimentacoes (id, tipo, material_nome, quantidade, unidade, motivo, data, usuario, created_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
                        (str(uuid.uuid4()), "sobra", m["nome"], quantidade, unidade, f"Sobra registrada: {descricao}", now_str, session.get("user_id"), now_iso)
                    )
                conn.commit()
            except Exception as e:
                try:
                    conn.rollback()
                except Exception:
                    pass
                flash('Erro ao registrar sobra: ' + str(e))
                return redirect(url_for("sobra_novo"))
            finally:
                try:
                    conn.close()
                except Exception:
                    pass
            flash(f'Sobra "{descricao}" registrada.')
            return redirect(url_for("sobras"))

        sobras_lista = carregar_json("sobras.json")
        sobras_lista.append({
            "id": str(uuid.uuid4()),
            "material_id": material_id or None,
            "descricao": descricao,
            "quantidade": quantidade,
            "unidade": unidade,
            "data": now.strftime("%d/%m/%Y"),
            "status": "Disponível",
        })
        if material_id and m:
            saldo_mat = float(m.get("quantidade") or 0)
            unidade_mat = (m.get("unidade") or "").lower()
            is_inteiro = unidade_mat.startswith("unid") or unidade_mat in ("unidades", "pares", "pacotes")
            nova_qtd = max(0, int(saldo_mat) - int(quantidade)) if is_inteiro else round(max(0.0, saldo_mat - float(quantidade)), 3)
            m["quantidade"] = nova_qtd
            salvar_materiais(materiais)
            registrar_movimentacao("sobra", quantidade, unidade, f"Sobra registrada: {descricao}", m.get("nome"))
        salvar_json("sobras.json", sobras_lista)
        flash(f'Sobra "{descricao}" registrada.')
        return redirect(url_for("sobras"))

    return render_template("sobra_form.html", materiais=materiais, unidades=UNIDADES)


@app.route("/sobras/<sobra_id>/reaproveitar", methods=["POST"])
@requires_permission('sobras', 'update')
def sobra_reaproveitar(sobra_id):
    now = agora()
    now_iso = now.isoformat()
    now_str = now.strftime("%d/%m/%Y %H:%M")
    usuario_id = session.get("user_id") if session else None

    if USE_SQLITE:
        init_db()
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        try:
            cur.execute("BEGIN IMMEDIATE")
            cur.execute("SELECT * FROM sobras WHERE id=?", (sobra_id,))
            s = cur.fetchone()
            if not s:
                conn.rollback()
                flash('Registro de sobra não encontrado.')
                return redirect(url_for('sobras'))
            if s["status"] != "Disponível":
                conn.rollback()
                flash('Esta sobra já foi processada ou não está disponível.')
                return redirect(url_for('sobras'))

            if s["material_id"]:
                cur.execute("SELECT quantidade, unidade, nome FROM materiais WHERE id=?", (s["material_id"],))
                mat = cur.fetchone()
                if not mat:
                    conn.rollback()
                    flash('Não foi possível reaproveitar: o material vinculado não existe no estoque.')
                    return redirect(url_for('sobras'))
                try:
                    unidade_mat = (mat["unidade"] or "").lower()
                    is_inteiro = unidade_mat.startswith("unid") or unidade_mat in ("unidades", "pares", "pacotes")
                    if is_inteiro:
                        nova = int(mat["quantidade"] or 0) + int(round(float(s["quantidade"] or 0)))
                    else:
                        nova = round(float(mat["quantidade"] or 0) + float(s["quantidade"] or 0), 3)
                except Exception:
                    conn.rollback()
                    flash('Quantidade inválida na sobra ou no material.')
                    return redirect(url_for('sobras'))
                cur.execute("UPDATE materiais SET quantidade=?, updated_at=? WHERE id=?", (nova, now_iso, s["material_id"]))
                cur.execute(
                    "INSERT INTO movimentacoes (id, tipo, material_nome, quantidade, unidade, motivo, data, usuario, created_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
                    (
                        str(uuid.uuid4()),
                        "reaproveitamento",
                        mat["nome"],
                        float(s["quantidade"] or 0),
                        s["unidade"],
                        f"Sobra reaproveitada de volta ao estoque (+{s['quantidade']} {s['unidade']})",
                        now_str,
                        usuario_id,
                        now_iso
                    )
                )

            cur.execute("UPDATE sobras SET status=?, updated_at=? WHERE id=?", ("Reaproveitado", now_iso, sobra_id))
            conn.commit()
            flash("Sobra reaproveitada com sucesso.")
        except Exception as e:
            try:
                conn.rollback()
            except Exception:
                pass
            flash('Erro ao reaproveitar sobra: ' + str(e))
        finally:
            try:
                conn.close()
            except Exception:
                pass
        return redirect(url_for('sobras'))

    sobras_lista = carregar_json("sobras.json")
    sobra = next((s for s in sobras_lista if s["id"] == sobra_id), None)
    if not sobra or sobra["status"] != "Disponível":
        flash("Sobra não disponível para reaproveitamento.")
        return redirect(url_for("sobras"))

    if sobra.get("material_id"):
        materiais = carregar_materiais()
        m = encontrar(materiais, sobra["material_id"])
        if not m:
            flash("Não foi possível reaproveitar: o material vinculado não existe no estoque.")
            return redirect(url_for("sobras"))
        unidade_mat = (m.get("unidade") or "").lower()
        is_inteiro = unidade_mat.startswith("unid") or unidade_mat in ("unidades", "pares", "pacotes")
        if is_inteiro:
            m["quantidade"] = int(m.get("quantidade") or 0) + int(round(float(sobra["quantidade"] or 0)))
        else:
            m["quantidade"] = round(float(m.get("quantidade") or 0) + float(sobra["quantidade"] or 0), 3)
        salvar_materiais(materiais)
        registrar_movimentacao("reaproveitamento", sobra["quantidade"], sobra["unidade"],
                                f"Sobra reaproveitada de volta ao estoque (+{sobra['quantidade']} {sobra['unidade']})", m["nome"])

    sobra["status"] = "Reaproveitado"
    salvar_json("sobras.json", sobras_lista)
    flash(f"{sobra['descricao']} reaproveitada com sucesso.")
    return redirect(url_for("sobras"))


@app.route("/sobras/<sobra_id>/descartar", methods=["POST"])
@requires_permission('sobras', 'update')
def sobra_descartar(sobra_id):
    if USE_SQLITE:
        init_db()
        conn = sqlite3.connect(DB_PATH)
        cur = conn.cursor()
        cur.execute("SELECT status FROM sobras WHERE id=?", (sobra_id,))
        r = cur.fetchone()
        if r and r[0] == "Disponível":
            cur.execute("UPDATE sobras SET status=?, updated_at=? WHERE id=?", ("Descartado", agora().isoformat(), sobra_id))
            conn.commit()
        conn.close()
        flash("Sobra marcada como descartada.")
        return redirect(url_for("sobras"))

    sobras_lista = carregar_json("sobras.json")
    sobra = next((s for s in sobras_lista if s["id"] == sobra_id), None)
    if sobra and sobra["status"] == "Disponível":
        sobra["status"] = "Descartado"
        salvar_json("sobras.json", sobras_lista)
        flash(f"{sobra['descricao']} marcada como descartada.")
    return redirect(url_for("sobras"))


@app.route("/sobras/<sobra_id>/excluir", methods=["POST"])
@requires_permission('sobras', 'delete')
def sobra_excluir(sobra_id):
    if USE_SQLITE:
        init_db()
        conn = sqlite3.connect(DB_PATH)
        cur = conn.cursor()
        cur.execute("DELETE FROM sobras WHERE id=?", (sobra_id,))
        conn.commit()
        conn.close()
        flash("Registro removido.")
        return redirect(url_for("sobras"))

    sobras_lista = carregar_json("sobras.json")
    sobras_lista = [s for s in sobras_lista if s["id"] != sobra_id]
    salvar_json("sobras.json", sobras_lista)
    flash("Registro removido.")
    return redirect(url_for("sobras"))


# ── Financeiro ─────────────────────────────────────────────────────────────────
@app.route("/financeiro")
@requires_permission('financeiro', 'read')
def financeiro():
    materiais = carregar_materiais()
    valor_estoque = round(sum(m["quantidade"] * m["custo"] for m in materiais), 2)

    if USE_SQLITE:
        init_db()
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        cur.execute("SELECT * FROM pedidos")
        pedidos_rows = cur.fetchall()
        pedidos_lista = [{
            "id": r["id"],
            "status": r["status"],
            "valor_total": r["valor_total"],
        } for r in pedidos_rows]
        receita_entregue = round(sum(p["valor_total"] for p in pedidos_lista if p["status"] == "Entregue"), 2)
        receita_prevista = round(sum(p["valor_total"] for p in pedidos_lista if p["status"] in ("Pendente", "Em produção", "Concluído")), 2)

        cur.execute("SELECT * FROM despesas ORDER BY created_at DESC")
        despesas_rows = cur.fetchall()
        despesas = [{"id": r["id"], "descricao": r["descricao"], "valor": r["valor"], "categoria": r["categoria"], "data": r["data"]} for r in despesas_rows]
        conn.close()
        total_despesas = round(sum(d["valor"] for d in despesas), 2)
        lucro = round(receita_entregue - total_despesas, 2)

        return render_template(
            "financeiro.html",
            valor_estoque=valor_estoque,
            receita_entregue=receita_entregue,
            receita_prevista=receita_prevista,
            despesas=list(reversed(despesas)),
            total_despesas=total_despesas,
            lucro=lucro,
            categorias_despesa=CATEGORIAS_DESPESA,
        )

    # legacy JSON path
    pedidos_lista = carregar_json("pedidos.json")
    receita_entregue = round(sum(p["valor_total"] for p in pedidos_lista if p["status"] == "Entregue"), 2)
    receita_prevista = round(
        sum(p["valor_total"] for p in pedidos_lista if p["status"] in ("Pendente", "Em produção", "Concluído")), 2
    )

    despesas = carregar_json("despesas.json")
    total_despesas = round(sum(d["valor"] for d in despesas), 2)
    lucro = round(receita_entregue - total_despesas, 2)

    return render_template(
        "financeiro.html",
        valor_estoque=valor_estoque,
        receita_entregue=receita_entregue,
        receita_prevista=receita_prevista,
        despesas=list(reversed(despesas)),
        total_despesas=total_despesas,
        lucro=lucro,
        categorias_despesa=CATEGORIAS_DESPESA,
    )


@app.route("/financeiro/despesa", methods=["POST"])
@requires_permission('financeiro', 'create')
def financeiro_despesa():
    descricao = request.form.get("descricao", "").strip()
    valor = parse_float_ptbr(request.form.get("valor", 0))
    categoria = request.form.get("categoria", "Outros")

    if not descricao or valor <= 0:
        flash("Informe descrição e valor válidos para a despesa.")
        return redirect(url_for("financeiro"))

    if USE_SQLITE:
        init_db()
        conn = sqlite3.connect(DB_PATH)
        cur = conn.cursor()
        now = agora().isoformat()
        cur.execute("INSERT INTO despesas (id,descricao,valor,categoria,data,created_at) VALUES (?,?,?,?,?,?)",
                    (str(uuid.uuid4()), descricao, float(valor), categoria, agora().strftime("%d/%m/%Y"), now))
        conn.commit()
        conn.close()
        flash(f'Despesa "{descricao}" registrada.')
        return redirect(url_for("financeiro"))

    despesas = carregar_json("despesas.json")
    despesas.append({
        "id": str(uuid.uuid4()),
        "descricao": descricao,
        "valor": valor,
        "categoria": categoria,
        "data": agora().strftime("%d/%m/%Y"),
    })
    salvar_json("despesas.json", despesas)
    flash(f'Despesa "{descricao}" registrada.')
    return redirect(url_for("financeiro"))


@app.route("/financeiro/despesa/<despesa_id>/excluir", methods=["POST"])
@requires_permission('financeiro', 'delete')
def financeiro_despesa_excluir(despesa_id):
    if USE_SQLITE:
        init_db()
        conn = sqlite3.connect(DB_PATH)
        cur = conn.cursor()
        cur.execute("DELETE FROM despesas WHERE id=?", (despesa_id,))
        conn.commit()
        conn.close()
        flash("Despesa removida.")
        return redirect(url_for("financeiro"))

    despesas = carregar_json("despesas.json")
    despesas = [d for d in despesas if d["id"] != despesa_id]
    salvar_json("despesas.json", despesas)
    flash("Despesa removida.")
    return redirect(url_for("financeiro"))


# ── Alertas e Relatórios ───────────────────────────────────────────────────────
@app.route("/alertas")
@app.route("/relatorios")
@requires_permission('relatorios', 'read')
def alertas():
    materiais = carregar_materiais()
    pedidos = carregar_pedidos()
    despesas = carregar_despesas()
    relatorios_personalizados = carregar_relatorios_customizados()

    baixo_estoque = sorted(
        [m for m in materiais if m.get("quantidade", 0) <= m.get("quantidade_minima", 0)],
        key=lambda m: m.get("quantidade", 0),
    )
    zerados = [m for m in baixo_estoque if m.get("quantidade", 0) == 0]
    valor_em_risco = round(sum(m.get("quantidade_minima", 0) * m.get("custo", 0) for m in baixo_estoque), 2)

    por_categoria = {}
    for m in materiais:
        c = por_categoria.setdefault(m.get("categoria", "Outros"), {"qtd_itens": 0, "valor": 0.0, "emoji": m.get("emoji", "📦")})
        c["qtd_itens"] += 1
        c["valor"] += m.get("quantidade", 0) * m.get("custo", 0)
    for c in por_categoria.values():
        c["valor"] = round(c["valor"], 2)

    movimentacoes = carregar_movimentacoes(30)

    # 1. Payload Gráfico: Estoque por Categoria
    chart_cat_labels = list(por_categoria.keys())
    chart_cat_values = [c["valor"] for c in por_categoria.values()]
    chart_cat_data = {
        "labels": chart_cat_labels,
        "datasets": [{
            "label": "Valor em Estoque (R$)",
            "data": chart_cat_values,
            "backgroundColor": PALETA_CORES_GRAFICO[:len(chart_cat_labels)],
            "borderWidth": 1.5,
            "borderColor": "#FFFFFF"
        }]
    }

    # 2. Payload Gráfico: Balanço Financeiro
    rec_entregue = round(sum(p.get("valor_total", 0) for p in pedidos if p.get("status") == "Entregue"), 2)
    rec_prevista = round(sum(p.get("valor_total", 0) for p in pedidos if p.get("status") in ("Pendente", "Em produção", "Concluído")), 2)
    tot_despesas = round(sum(d.get("valor", 0) for d in despesas), 2)
    lucro = round(rec_entregue - tot_despesas, 2)

    chart_fin_data = {
        "labels": ["Receita Recebida", "Receita Prevista", "Despesas Totais", "Lucro Realizado"],
        "datasets": [{
            "label": "Valor (R$)",
            "data": [rec_entregue, rec_prevista, tot_despesas, max(0, lucro)],
            "backgroundColor": ["#2E7D32", "#C88242", "#C62828", "#7C3D12"],
            "borderColor": ["#1B5E20", "#A05A18", "#B71C1C", "#5C2D0E"],
            "borderWidth": 1.5
        }]
    }

    # 3. Payload Gráfico: Status dos Pedidos
    status_counts = {}
    for p in pedidos:
        st = p.get("status", "Pendente")
        status_counts[st] = status_counts.get(st, 0) + 1
    
    chart_ped_labels = list(status_counts.keys())
    chart_ped_data = {
        "labels": chart_ped_labels,
        "datasets": [{
            "label": "Quantidade de Pedidos",
            "data": [status_counts[k] for k in chart_ped_labels],
            "backgroundColor": PALETA_CORES_GRAFICO[:len(chart_ped_labels)],
            "borderWidth": 1.5,
            "borderColor": "#FFFFFF"
        }]
    }

    return render_template(
        "alertas.html",
        baixo_estoque=baixo_estoque,
        zerados=zerados,
        valor_em_risco=valor_em_risco,
        por_categoria=por_categoria,
        movimentacoes=movimentacoes,
        total_materiais=len(materiais),
        relatorios_personalizados=relatorios_personalizados,
        chart_cat_data=json.dumps(chart_cat_data, ensure_ascii=False),
        chart_fin_data=json.dumps(chart_fin_data, ensure_ascii=False),
        chart_ped_data=json.dumps(chart_ped_data, ensure_ascii=False),
    )


# Criar Relatório Personalizado
@app.route("/relatorios/novo", methods=["GET", "POST"])
@requires_permission('relatorios', 'create')
def relatorio_novo():
    if request.method == "POST":
        titulo = request.form.get("titulo", "").strip()
        tipo = request.form.get("tipo", "estoque")
        tipo_grafico = request.form.get("tipo_grafico", "bar")
        categoria_filtro = request.form.get("categoria_filtro", "").strip()
        status_filtro = request.form.get("status_filtro", "").strip()
        apenas_criticos = bool(request.form.get("apenas_criticos"))
        observacoes = request.form.get("observacoes", "").strip()

        if not titulo:
            flash("Informe um título para o relatório.")
            return redirect(url_for("relatorio_novo"))

        criado_por = g.user.get("nome") or g.user.get("username") or "Usuário"
        novo_rel = {
            "titulo": titulo,
            "tipo": tipo,
            "tipo_grafico": tipo_grafico,
            "categoria_filtro": categoria_filtro,
            "status_filtro": status_filtro,
            "apenas_criticos": apenas_criticos,
            "observacoes": observacoes,
            "criado_por": criado_por,
        }
        salvo = salvar_relatorio_customizado(novo_rel)
        flash(f'Relatório "{titulo}" criado com sucesso!')
        return redirect(url_for("relatorio_detalhe", relatorio_id=salvo["id"]))

    return render_template(
        "relatorio_form.html",
        relatorio=None,
        categorias=CATEGORIAS,
        status_pedido=STATUS_PEDIDO,
    )


# Visualizar Relatório Personalizado
@app.route("/relatorios/<relatorio_id>")
@requires_permission('relatorios', 'read')
def relatorio_detalhe(relatorio_id):
    relatorio = encontrar_relatorio_por_id(relatorio_id)
    if not relatorio:
        flash("Relatório não encontrado.")
        return redirect(url_for("alertas"))

    dados = gerar_dados_relatorio(relatorio)
    return render_template(
        "relatorio_detalhe.html",
        relatorio=relatorio,
        dados=dados,
        chart_payload=json.dumps(dados["chart_data"], ensure_ascii=False),
    )


# Editar Relatório Personalizado
@app.route("/relatorios/<relatorio_id>/editar", methods=["GET", "POST"])
@requires_permission('relatorios', 'update')
def relatorio_editar(relatorio_id):
    relatorio = encontrar_relatorio_por_id(relatorio_id)
    if not relatorio:
        flash("Relatório não encontrado.")
        return redirect(url_for("alertas"))

    if request.method == "POST":
        titulo = request.form.get("titulo", "").strip()
        tipo = request.form.get("tipo", "estoque")
        tipo_grafico = request.form.get("tipo_grafico", "bar")
        categoria_filtro = request.form.get("categoria_filtro", "").strip()
        status_filtro = request.form.get("status_filtro", "").strip()
        apenas_criticos = bool(request.form.get("apenas_criticos"))
        observacoes = request.form.get("observacoes", "").strip()

        if not titulo:
            flash("Informe um título para o relatório.")
            return redirect(url_for("relatorio_editar", relatorio_id=relatorio_id))

        relatorio["titulo"] = titulo
        relatorio["tipo"] = tipo
        relatorio["tipo_grafico"] = tipo_grafico
        relatorio["categoria_filtro"] = categoria_filtro
        relatorio["status_filtro"] = status_filtro
        relatorio["apenas_criticos"] = apenas_criticos
        relatorio["observacoes"] = observacoes

        salvar_relatorio_customizado(relatorio)
        flash(f'Relatório "{titulo}" atualizado!')
        return redirect(url_for("relatorio_detalhe", relatorio_id=relatorio_id))

    return render_template(
        "relatorio_form.html",
        relatorio=relatorio,
        categorias=CATEGORIAS,
        status_pedido=STATUS_PEDIDO,
    )


# Excluir Relatório Personalizado
@app.route("/relatorios/<relatorio_id>/excluir", methods=["POST"])
@requires_permission('relatorios', 'delete')
def relatorio_excluir(relatorio_id):
    relatorio = encontrar_relatorio_por_id(relatorio_id)
    if not relatorio:
        flash("Relatório não encontrado.")
        return redirect(url_for("alertas"))

    titulo = relatorio.get("titulo", "Relatório")
    excluir_relatorio_customizado(relatorio_id)
    flash(f'Relatório "{titulo}" excluído.')
    return redirect(url_for("alertas"))


# Serve uploaded files
@app.route('/uploads/<path:filename>')
def uploaded_file(filename):
    uploads_dir = os.path.join(DATA_DIR, 'uploads')
    return send_from_directory(uploads_dir, filename)


# Exportar todos os dados (backup)
@app.route("/exportar")
@requires_permission('relatorios', 'read')
def exportar_tudo():
    colecoes = [
        "materiais.json",
        "produtos.json",
        "pedidos.json",
        "movimentacoes.json",
        "sobras.json",
        "despesas.json",
    ]
    tudo = {}
    for c in colecoes:
        tudo[os.path.splitext(c)[0]] = carregar_json(c)
    resp = make_response(json.dumps(tudo, ensure_ascii=False, indent=2))
    resp.headers["Content-Type"] = "application/json; charset=utf-8"
    resp.headers["Content-Disposition"] = "attachment; filename=export_all.json"
    return resp


# Exportar relatório financeiro em PDF com design e paleta do site
@app.route('/exportar/pdf')
@requires_permission('relatorios', 'read')
def exportar_financeiro_pdf():
    try:
        from io import BytesIO
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import mm
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table as RLTable, TableStyle
        )
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.pdfgen import canvas
    except ImportError as e:
        return str(e), 400

    class NumberedCanvas(canvas.Canvas):
        def __init__(self, *args, **kwargs):
            super().__init__(*args, **kwargs)
            self._saved_page_states = []

        def showPage(self):
            self._saved_page_states.append(dict(self.__dict__))
            self._startPage()

        def save(self):
            num_pages = len(self._saved_page_states)
            for state in self._saved_page_states:
                self.__dict__.update(state)
                self.draw_page_decorations(num_pages)
                super().showPage()
            super().save()

        def draw_page_decorations(self, page_count):
            self.saveState()
            width, height = A4
            # Top banner
            self.setFillColor(colors.HexColor("#7C3D12"))
            self.rect(0, height - 16 * mm, width, 16 * mm, stroke=0, fill=1)

            # Header text
            self.setFillColor(colors.HexColor("#FDE9C2"))
            self.setFont("Helvetica-Bold", 11)
            self.drawString(15 * mm, height - 10.5 * mm, "ATELIE HAITI  -  GESTAO ARTESANAL")

            self.setFont("Helvetica", 8.5)
            self.setFillColor(colors.white)
            self.drawRightString(width - 15 * mm, height - 10.5 * mm, f"Emissao: {agora().strftime('%d/%m/%Y %H:%M')}")

            # Footer line
            self.setStrokeColor(colors.HexColor("#E2D2BC"))
            self.setLineWidth(0.8)
            self.line(15 * mm, 14 * mm, width - 15 * mm, 14 * mm)

            # Footer text
            self.setFont("Helvetica-Oblique", 8)
            self.setFillColor(colors.HexColor("#7A6B63"))
            self.drawString(15 * mm, 9 * mm, "Conectados pela Comunidade - Favela do Haiti, SP")
            self.drawRightString(width - 15 * mm, 9 * mm, f"Pagina {self._pageNumber} de {page_count}")
            self.restoreState()

    materiais = carregar_materiais()
    pedidos = carregar_pedidos()
    despesas = carregar_despesas()

    valor_estoque = round(sum(m.get("quantidade", 0) * m.get("custo", 0) for m in materiais), 2)
    receita_entregue = round(sum(p.get("valor_total", 0) for p in pedidos if p.get("status") == "Entregue"), 2)
    receita_prevista = round(sum(p.get("valor_total", 0) for p in pedidos if p.get("status") in ("Pendente", "Em produção", "Concluído")), 2)
    total_despesas = round(sum(d.get("valor", 0) for d in despesas), 2)
    lucro = round(receita_entregue - total_despesas, 2)

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        leftMargin=15 * mm,
        rightMargin=15 * mm,
        topMargin=22 * mm,
        bottomMargin=18 * mm
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=16,
        leading=20,
        textColor=colors.HexColor("#7C3D12"),
        spaceAfter=3
    )
    subtitle_style = ParagraphStyle(
        'DocSub',
        parent=styles['Normal'],
        fontName='Helvetica-Oblique',
        fontSize=9.5,
        leading=13,
        textColor=colors.HexColor("#7A6B63"),
        spaceAfter=10
    )
    h2_style = ParagraphStyle(
        'H2',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=11,
        leading=15,
        textColor=colors.HexColor("#5C2D0E"),
        spaceBefore=8,
        spaceAfter=5
    )
    cell_style = ParagraphStyle(
        'Cell',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8,
        leading=10,
        textColor=colors.HexColor("#2C1810")
    )
    cell_bold = ParagraphStyle(
        'CellBold',
        parent=cell_style,
        fontName='Helvetica-Bold'
    )
    cell_right = ParagraphStyle(
        'CellRight',
        parent=cell_style,
        alignment=2
    )
    cell_right_bold = ParagraphStyle(
        'CellRightBold',
        parent=cell_bold,
        alignment=2
    )

    elements = []
    elements.append(Paragraph("Relatório Financeiro &amp; Balanço Executivo", title_style))
    elements.append(Paragraph("Visão consolidada de estoque, faturamento de pedidos, despesas e margem operacional.", subtitle_style))

    # 1. KPI Cards Grid Table
    lucro_color = '#2E7D32' if lucro >= 0 else '#C62828'
    kpi_data = [
        [
            Paragraph(f"<b>Valor em Estoque</b><br/><font size=12 color='#7C3D12'><b>R$ {formatar_reais(valor_estoque)}</b></font><br/><font size=7 color='#7A6B63'>Total em insumos</font>", cell_style),
            Paragraph(f"<b>Receita Recebida</b><br/><font size=12 color='#2E7D32'><b>R$ {formatar_reais(receita_entregue)}</b></font><br/><font size=7 color='#7A6B63'>Pedidos entregues</font>", cell_style),
            Paragraph(f"<b>Receita Prevista</b><br/><font size=12 color='#C88242'><b>R$ {formatar_reais(receita_prevista)}</b></font><br/><font size=7 color='#7A6B63'>Pedidos em andamento</font>", cell_style),
        ],
        [
            Paragraph(f"<b>Despesas Totais</b><br/><font size=12 color='#C62828'><b>R$ {formatar_reais(total_despesas)}</b></font><br/><font size=7 color='#7A6B63'>Custos operacionais</font>", cell_style),
            Paragraph(f"<b>Lucro Realizado</b><br/><font size=12 color='{lucro_color}'><b>R$ {formatar_reais(lucro)}</b></font><br/><font size=7 color='#7A6B63'>Receita − Despesas</font>", cell_style),
            Paragraph(f"<b>Total de Materiais</b><br/><font size=12 color='#5C2D0E'><b>{len(materiais)} itens</b></font><br/><font size=7 color='#7A6B63'>Cadastrados no estoque</font>", cell_style),
        ]
    ]
    t_kpi = RLTable(kpi_data, colWidths=[60 * mm, 60 * mm, 60 * mm])
    t_kpi.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor("#FDF8F0")),
        ('BOX', (0,0), (-1,-1), 1, colors.HexColor("#E2D2BC")),
        ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor("#E2D2BC")),
        ('TOPPADDING', (0,0), (-1,-1), 5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ('LEFTPADDING', (0,0), (-1,-1), 8),
        ('RIGHTPADDING', (0,0), (-1,-1), 8),
    ]))
    elements.append(t_kpi)
    elements.append(Spacer(1, 10))

    # 2. Despesas Registradas
    elements.append(Paragraph("Despesas Recentes", h2_style))
    desp_hdr = [
        Paragraph("<font color='white'><b>Data</b></font>", cell_bold),
        Paragraph("<font color='white'><b>Descrição</b></font>", cell_bold),
        Paragraph("<font color='white'><b>Categoria</b></font>", cell_bold),
        Paragraph("<font color='white'><b>Valor (R$)</b></font>", cell_right_bold),
    ]
    desp_table_data = [desp_hdr]
    if despesas:
        for d in despesas[:12]:
            desp_table_data.append([
                Paragraph(str(d.get("data", "")), cell_style),
                Paragraph(str(d.get("descricao", "")), cell_style),
                Paragraph(str(d.get("categoria", "Outros")), cell_style),
                Paragraph(f"R$ {formatar_reais(float(d.get('valor', 0)))}", cell_right),
            ])
    else:
        desp_table_data.append([Paragraph("Nenhuma despesa registrada.", cell_style), "", "", ""])

    t_desp = RLTable(desp_table_data, colWidths=[25 * mm, 80 * mm, 40 * mm, 35 * mm])
    t_desp.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#7C3D12")),
        ('BOTTOMPADDING', (0,0), (-1,-1), 3),
        ('TOPPADDING', (0,0), (-1,-1), 3),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.HexColor("#FFFFFF"), colors.HexColor("#FFFDF9")]),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#E2D2BC")),
    ]))
    elements.append(t_desp)
    elements.append(Spacer(1, 10))

    # 3. Pedidos dos Clientes
    elements.append(Paragraph("Pedidos dos Clientes", h2_style))
    ped_hdr = [
        Paragraph("<font color='white'><b>Cliente</b></font>", cell_bold),
        Paragraph("<font color='white'><b>Produto</b></font>", cell_bold),
        Paragraph("<font color='white'><b>Qtd</b></font>", cell_bold),
        Paragraph("<font color='white'><b>Status</b></font>", cell_bold),
        Paragraph("<font color='white'><b>Valor Total</b></font>", cell_right_bold),
    ]
    ped_table_data = [ped_hdr]
    if pedidos:
        for p in pedidos[:12]:
            st = p.get("status", "Pendente")
            ped_table_data.append([
                Paragraph(str(p.get("cliente", "")), cell_style),
                Paragraph(str(p.get("produto_nome", "")), cell_style),
                Paragraph(str(p.get("quantidade", 1)), cell_style),
                Paragraph(st, cell_style),
                Paragraph(f"R$ {formatar_reais(float(p.get('valor_total', 0)))}", cell_right),
            ])
    else:
        ped_table_data.append([Paragraph("Nenhum pedido registrado.", cell_style), "", "", "", ""])

    t_ped = RLTable(ped_table_data, colWidths=[45 * mm, 50 * mm, 15 * mm, 35 * mm, 35 * mm])
    t_ped.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#7C3D12")),
        ('BOTTOMPADDING', (0,0), (-1,-1), 3),
        ('TOPPADDING', (0,0), (-1,-1), 3),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.HexColor("#FFFFFF"), colors.HexColor("#FFFDF9")]),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#E2D2BC")),
    ]))
    elements.append(t_ped)
    elements.append(Spacer(1, 10))

    # 4. Alertas de Estoque Baixo
    baixo_estoque = [m for m in materiais if m.get("quantidade", 0) <= m.get("quantidade_minima", 0)]
    if baixo_estoque:
        elements.append(Paragraph("Materiais em Nível Crítico de Estoque", h2_style))
        crit_hdr = [
            Paragraph("<font color='white'><b>Material</b></font>", cell_bold),
            Paragraph("<font color='white'><b>Categoria</b></font>", cell_bold),
            Paragraph("<font color='white'><b>Qtd Atual</b></font>", cell_bold),
            Paragraph("<font color='white'><b>Qtd Mínima</b></font>", cell_bold),
            Paragraph("<font color='white'><b>Custo Reposição</b></font>", cell_right_bold),
        ]
        crit_table_data = [crit_hdr]
        for m in baixo_estoque:
            custo_rep = (m.get("quantidade_minima", 0) - m.get("quantidade", 0)) * m.get("custo", 0)
            if custo_rep < 0:
                custo_rep = 0
            crit_table_data.append([
                Paragraph(f"{m.get('nome','')}", cell_style),
                Paragraph(str(m.get('categoria','')), cell_style),
                Paragraph(f"<font color='#C62828'><b>{m.get('quantidade',0)} {m.get('unidade','')}</b></font>", cell_style),
                Paragraph(f"{m.get('quantidade_minima',0)} {m.get('unidade','')}", cell_style),
                Paragraph(f"R$ {formatar_reais(custo_rep)}", cell_right),
            ])
        t_crit = RLTable(crit_table_data, colWidths=[55 * mm, 35 * mm, 30 * mm, 25 * mm, 35 * mm])
        t_crit.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#C62828")),
            ('BOTTOMPADDING', (0,0), (-1,-1), 3),
            ('TOPPADDING', (0,0), (-1,-1), 3),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.HexColor("#FFF8F8"), colors.HexColor("#FFFFFF")]),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#E2D2BC")),
        ]))
        elements.append(t_crit)

    doc.build(elements, canvasmaker=NumberedCanvas)
    buf.seek(0)
    return send_file(buf, mimetype='application/pdf', as_attachment=True, download_name='relatorio_financeiro_atelie.pdf')


# Exportar todos os dados em Excel (XLSX) estruturado em Tabelas
@app.route('/exportar/xlsx')
@requires_permission('relatorios', 'read')
def exportar_tudo_xlsx():
    try:
        from io import BytesIO
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.table import Table, TableStyleInfo
    except ImportError as e:
        return str(e), 400

    wb = Workbook()

    header_fill = PatternFill(start_color="7C3D12", end_color="7C3D12", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Segoe UI", size=10)
    data_bold = Font(name="Segoe UI", size=10, bold=True)

    thin_border = Border(
        left=Side(style='thin', color='E2D2BC'),
        right=Side(style='thin', color='E2D2BC'),
        top=Side(style='thin', color='E2D2BC'),
        bottom=Side(style='thin', color='E2D2BC')
    )
    alt_fill = PatternFill(start_color="FDF8F0", end_color="FDF8F0", fill_type="solid")
    warning_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
    warning_font = Font(name="Segoe UI", size=10, bold=True, color="C62828")

    # 1. ABA RESUMO GERAL
    ws_resumo = wb.active
    ws_resumo.title = "Resumo Geral"
    ws_resumo.views.sheetView[0].showGridLines = True

    ws_resumo.merge_cells("A1:E1")
    ws_resumo["A1"] = "ATELIÊ HAITI — RELATÓRIO GERAL E BALANÇO"
    ws_resumo["A1"].font = Font(name="Segoe UI", size=14, bold=True, color="FFFFFF")
    ws_resumo["A1"].fill = header_fill
    ws_resumo["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_resumo.row_dimensions[1].height = 34

    ws_resumo["A2"] = f"Relatório gerado em: {agora().strftime('%d/%m/%Y %H:%M:%S')}"
    ws_resumo["A2"].font = Font(name="Segoe UI", size=9.5, italic=True, color="7A6B63")
    ws_resumo.row_dimensions[2].height = 18

    materiais = carregar_materiais()
    pedidos = carregar_pedidos()
    despesas = carregar_despesas()
    produtos = carregar_produtos()
    sobras = carregar_sobras()
    movimentacoes = carregar_movimentacoes(200)

    valor_estoque = round(sum(m.get("quantidade", 0) * m.get("custo", 0) for m in materiais), 2)
    receita_entregue = round(sum(p.get("valor_total", 0) for p in pedidos if p.get("status") == "Entregue"), 2)
    receita_prevista = round(sum(p.get("valor_total", 0) for p in pedidos if p.get("status") in ("Pendente", "Em produção", "Concluído")), 2)
    total_despesas = round(sum(d.get("valor", 0) for d in despesas), 2)
    lucro = round(receita_entregue - total_despesas, 2)

    resumo_kpis = [
        ("Indicador Financeiro / Operacional", "Valor Consolidado", "Status / Detalhe"),
        ("Valor Total em Estoque (Insumos)", valor_estoque, f"{len(materiais)} materiais cadastrados"),
        ("Receita Confirmada (Pedidos Entregues)", receita_entregue, f"{len([p for p in pedidos if p.get('status') == 'Entregue'])} pedidos finalizados"),
        ("Receita Prevista (Em Produção/Pendentes)", receita_prevista, f"{len([p for p in pedidos if p.get('status') != 'Entregue'])} pedidos em andamento"),
        ("Despesas Totais Registradas", total_despesas, f"{len(despesas)} lançamentos"),
        ("Lucro Realizado (Receita − Despesas)", lucro, "Saldo operacional líquido"),
    ]

    start_row = 4
    for r_idx, row_data in enumerate(resumo_kpis, start=start_row):
        for c_idx, val in enumerate(row_data, start=1):
            cell = ws_resumo.cell(row=r_idx, column=c_idx, value=val)
            cell.border = thin_border
            if r_idx == start_row:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center" if c_idx > 1 else "left", vertical="center")
            else:
                if c_idx == 2 and isinstance(val, (int, float)):
                    cell.number_format = 'R$ #,##0.00'
                    cell.font = data_bold
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    cell.font = data_font
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                if r_idx % 2 == 0:
                    cell.fill = alt_fill
        ws_resumo.row_dimensions[r_idx].height = 22

    # 2. ABA ESTOQUE DE MATERIAIS
    ws_mat = wb.create_sheet(title="Estoque de Materiais")
    ws_mat.views.sheetView[0].showGridLines = True
    mat_headers = ["ID", "Material", "Categoria", "Qtd em Estoque", "Unidade", "Qtd Mínima", "Custo Unitário", "Valor Total", "GTIN", "Status"]
    ws_mat.append(mat_headers)
    ws_mat.row_dimensions[1].height = 24

    for r_idx, m in enumerate(materiais, start=2):
        qtd = float(m.get("quantidade", 0))
        qtd_min = float(m.get("quantidade_minima", 0))
        custo = float(m.get("custo", 0))
        val_tot = qtd * custo
        is_critico = (qtd <= qtd_min)
        status_txt = "Crítico" if is_critico else "Normal"

        row = [
            m.get("id", ""),
            f"{m.get('nome', '')}",
            m.get("categoria", ""),
            qtd,
            m.get("unidade", ""),
            qtd_min,
            custo,
            val_tot,
            m.get("gtin", "") or "-",
            status_txt
        ]
        ws_mat.append(row)
        ws_mat.row_dimensions[r_idx].height = 20

        for c_idx in range(1, len(mat_headers) + 1):
            cell = ws_mat.cell(row=r_idx, column=c_idx)
            cell.border = thin_border
            cell.font = data_font
            if c_idx in (4, 6):
                cell.number_format = '#,##0.00'
            elif c_idx in (7, 8):
                cell.number_format = 'R$ #,##0.00'

            if is_critico:
                cell.fill = warning_fill
                if c_idx in (4, 10):
                    cell.font = warning_font
            elif r_idx % 2 == 0:
                cell.fill = alt_fill

    for c_idx in range(1, len(mat_headers) + 1):
        cell = ws_mat.cell(row=1, column=c_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center" if c_idx not in (2, 3) else "left", vertical="center")
        cell.border = thin_border

    if materiais:
        tab_mat = Table(displayName="TabelaEstoque", ref=f"A1:{get_column_letter(len(mat_headers))}{len(materiais) + 1}")
        tab_mat.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        ws_mat.add_table(tab_mat)

    # 3. ABA PRODUTOS
    ws_prod = wb.create_sheet(title="Produtos")
    ws_prod.views.sheetView[0].showGridLines = True
    prod_headers = ["ID", "Produto", "Preço de Venda", "Qtd Insumos Receita"]
    ws_prod.append(prod_headers)
    ws_prod.row_dimensions[1].height = 24

    for r_idx, pr in enumerate(produtos, start=2):
        receita_len = len(pr.get("receita") or [])
        row = [
            pr.get("id", ""),
            pr.get("nome", ""),
            float(pr.get("preco_venda", 0)),
            receita_len
        ]
        ws_prod.append(row)
        ws_prod.row_dimensions[r_idx].height = 20
        for c_idx in range(1, len(prod_headers) + 1):
            cell = ws_prod.cell(row=r_idx, column=c_idx)
            cell.border = thin_border
            cell.font = data_font
            if c_idx == 3:
                cell.number_format = 'R$ #,##0.00'
            elif c_idx == 4:
                cell.number_format = '#,##0'

    for c_idx in range(1, len(prod_headers) + 1):
        cell = ws_prod.cell(row=1, column=c_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center" if c_idx != 2 else "left", vertical="center")
        cell.border = thin_border

    if produtos:
        tab_prod = Table(displayName="TabelaProdutos", ref=f"A1:{get_column_letter(len(prod_headers))}{len(produtos) + 1}")
        tab_prod.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        ws_prod.add_table(tab_prod)

    # 4. ABA PEDIDOS
    ws_ped = wb.create_sheet(title="Pedidos")
    ws_ped.views.sheetView[0].showGridLines = True
    ped_headers = ["ID", "Cliente", "Produto", "Quantidade", "Valor Unitário", "Valor Total", "Status", "Data Entrega", "Criado em"]
    ws_ped.append(ped_headers)
    ws_ped.row_dimensions[1].height = 24

    for r_idx, p in enumerate(pedidos, start=2):
        qtd = float(p.get("quantidade", 1))
        val_tot = float(p.get("valor_total", 0))
        val_un = round(val_tot / qtd, 2) if qtd > 0 else 0
        row = [
            p.get("id", ""),
            p.get("cliente", ""),
            p.get("produto_nome", ""),
            qtd,
            val_un,
            val_tot,
            p.get("status", "Pendente"),
            p.get("data_entrega", "") or "-",
            p.get("created_at", "")[:16] if p.get("created_at") else "-"
        ]
        ws_ped.append(row)
        ws_ped.row_dimensions[r_idx].height = 20
        for c_idx in range(1, len(ped_headers) + 1):
            cell = ws_ped.cell(row=r_idx, column=c_idx)
            cell.border = thin_border
            cell.font = data_font
            if c_idx in (5, 6):
                cell.number_format = 'R$ #,##0.00'
            elif c_idx == 4:
                cell.number_format = '#,##0'

    for c_idx in range(1, len(ped_headers) + 1):
        cell = ws_ped.cell(row=1, column=c_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center" if c_idx not in (2, 3) else "left", vertical="center")
        cell.border = thin_border

    if pedidos:
        tab_ped = Table(displayName="TabelaPedidos", ref=f"A1:{get_column_letter(len(ped_headers))}{len(pedidos) + 1}")
        tab_ped.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        ws_ped.add_table(tab_ped)

    # 5. ABA DESPESAS
    ws_desp = wb.create_sheet(title="Despesas")
    ws_desp.views.sheetView[0].showGridLines = True
    desp_headers = ["ID", "Descrição", "Categoria", "Valor", "Data", "Criado em"]
    ws_desp.append(desp_headers)
    ws_desp.row_dimensions[1].height = 24

    for r_idx, d in enumerate(despesas, start=2):
        row = [
            d.get("id", ""),
            d.get("descricao", ""),
            d.get("categoria", "Outros"),
            float(d.get("valor", 0)),
            d.get("data", ""),
            d.get("created_at", "")[:16] if d.get("created_at") else "-"
        ]
        ws_desp.append(row)
        ws_desp.row_dimensions[r_idx].height = 20
        for c_idx in range(1, len(desp_headers) + 1):
            cell = ws_desp.cell(row=r_idx, column=c_idx)
            cell.border = thin_border
            cell.font = data_font
            if c_idx == 4:
                cell.number_format = 'R$ #,##0.00'
                cell.font = data_bold

    for c_idx in range(1, len(desp_headers) + 1):
        cell = ws_desp.cell(row=1, column=c_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center" if c_idx != 2 else "left", vertical="center")
        cell.border = thin_border

    if despesas:
        tab_desp = Table(displayName="TabelaDespesas", ref=f"A1:{get_column_letter(len(desp_headers))}{len(despesas) + 1}")
        tab_desp.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        ws_desp.add_table(tab_desp)

    # 6. ABA SOBRAS
    ws_sob = wb.create_sheet(title="Sobras")
    ws_sob.views.sheetView[0].showGridLines = True
    sob_headers = ["ID", "Descrição", "Quantidade", "Unidade", "Data", "Status"]
    ws_sob.append(sob_headers)
    ws_sob.row_dimensions[1].height = 24

    for r_idx, s in enumerate(sobras, start=2):
        row = [
            s.get("id", ""),
            s.get("descricao", ""),
            float(s.get("quantidade", 0)),
            s.get("unidade", ""),
            s.get("data", ""),
            s.get("status", "Disponível")
        ]
        ws_sob.append(row)
        ws_sob.row_dimensions[r_idx].height = 20
        for c_idx in range(1, len(sob_headers) + 1):
            cell = ws_sob.cell(row=r_idx, column=c_idx)
            cell.border = thin_border
            cell.font = data_font
            if c_idx == 3:
                cell.number_format = '#,##0.00'

    for c_idx in range(1, len(sob_headers) + 1):
        cell = ws_sob.cell(row=1, column=c_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center" if c_idx != 2 else "left", vertical="center")
        cell.border = thin_border

    if sobras:
        tab_sob = Table(displayName="TabelaSobras", ref=f"A1:{get_column_letter(len(sob_headers))}{len(sobras) + 1}")
        tab_sob.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        ws_sob.add_table(tab_sob)

    # Auto-ajustar largura das colunas
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                if cell.number_format and 'R$' in cell.number_format:
                    val_str = f"R$ {val_str},00"
                max_len = max(max_len, len(val_str))
            sheet.column_dimensions[col_letter].width = max(max_len + 4, 12)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name='export_atelie_haiti.xlsx')


# Fallback — mantém a navegação de pé para qualquer rota que ainda não exista.
@app.route("/<pagina>")
def em_construcao(pagina):
    return redirect(url_for("home"))


if __name__ == "__main__":
    app.run(debug=True)


